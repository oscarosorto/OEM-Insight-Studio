
import streamlit as st
import numpy as np
import openpyxl
import matplotlib.pyplot as plt
from matplotlib.colors import LinearSegmentedColormap, Normalize
from matplotlib.gridspec import GridSpec, GridSpecFromSubplotSpec
from io import BytesIO
from datetime import date
from pathlib import Path

st.set_page_config(page_title="OEM Insight Studio", page_icon="📊", layout="wide")

st.markdown("""
<style>
.stApp {background:#f3f6fb;}
.header-box {
    background: white;
    padding: 18px 22px;
    border-radius: 18px;
    border: 1px solid #d9e2ef;
    box-shadow: 0 4px 14px rgba(15,23,42,.06);
}
.title {color:#0b2545;font-size:32px;font-weight:800;}
.subtitle {color:#31517a;font-size:16px;}
</style>
""", unsafe_allow_html=True)

# ==================================================
# LECTURA DE DATOS
# ==================================================
def load_wb(uploaded_file):
    uploaded_file.seek(0)
    return openpyxl.load_workbook(uploaded_file, data_only=True)

def get_sheets(uploaded_file):
    return load_wb(uploaded_file).sheetnames

def read_template(uploaded_file, sheet_name, upper=True):
    wb = load_wb(uploaded_file)
    ws = wb[sheet_name]
    rows = range(14, 19)
    cols = [7, 8, 9] if upper else [18, 19, 20]
    arr = []
    for r in rows:
        row = []
        for c in cols:
            val = ws.cell(r, c).value
            if val is None:
                raise ValueError(f"Celda vacía en hoja {sheet_name}, fila {r}, columna {c}")
            row.append(float(val))
        arr.append(row)
    return np.array(arr, dtype=float)

def read_manual(uploaded_file, sheet_name, start_row, start_col):
    wb = load_wb(uploaded_file)
    ws = wb[sheet_name]
    arr = []
    for r in range(start_row, start_row + 5):
        row = []
        for c in range(start_col, start_col + 3):
            val = ws.cell(r, c).value
            row.append(float(val))
        arr.append(row)
    return np.array(arr, dtype=float)

# ==================================================
# CÁLCULOS Y GRÁFICOS
# ==================================================
def condition(arr):
    mn = float(np.min(arr))
    rg = float(np.max(arr)-np.min(arr))
    if mn < 0.85 or rg > 0.50:
        return "REVISAR", "#e10600"
    elif mn < 1.05 or rg > 0.28:
        return "MONITOREAR", "#f5b400"
    else:
        return "NORMAL", "#228b22"

def interp_surface(arr, nx=220, ny=150):
    x = np.arange(1, 6)
    y = np.arange(1, 4)
    xi = np.linspace(1, 5, nx)
    temp = np.array([np.interp(xi, x, arr[:, j]) for j in range(3)])
    yi = np.linspace(1, 3, ny)
    zi = np.array([np.interp(yi, y, temp[:, i]) for i in range(nx)]).T
    X, Y = np.meshgrid(xi, yi)
    return X, Y, zi

def centroid(arr):
    wear = arr.max() - arr
    total = wear.sum()
    Xp, Yp = np.meshgrid(np.arange(1,6), np.arange(1,4), indexing="ij")
    if total <= 0:
        return 3, 2
    return (Xp * wear).sum()/total, (Yp * wear).sum()/total

def add_panel(fig, outer_spec, title, arr, cmap, norm, panel_color):
    gs = GridSpecFromSubplotSpec(
        1, 3, subplot_spec=outer_spec,
        width_ratios=[4.2, 1.45, 1.15],
        wspace=0.12
    )
    ax_map = fig.add_subplot(gs[0,0])
    ax_tbl = fig.add_subplot(gs[0,1])
    ax_kpi = fig.add_subplot(gs[0,2])

    # Map
    X, Y, Z = interp_surface(arr)
    ax_map.imshow(
        Z.T, origin="lower", extent=[1,5,1,3],
        cmap=cmap, norm=norm, interpolation="bicubic",
        aspect="auto"
    )
    levels = np.linspace(norm.vmin, norm.vmax, 5)
    cs = ax_map.contour(X, Y, Z, levels=levels, colors="#0b2545", linewidths=0.45, alpha=0.35)
    ax_map.clabel(cs, fontsize=6, inline=True, fmt="%.2f")

    for i in range(5):
        for j in range(3):
            ax_map.scatter(i+1, j+1, s=24, facecolor="white", edgecolor="black", linewidth=0.8, zorder=5)

    mi = np.unravel_index(np.argmin(arr), arr.shape)
    ax_map.scatter(mi[0]+1, mi[1]+1, s=70, marker="s", facecolor="white", edgecolor="black", linewidth=1.3, zorder=6)

    cx, cy = centroid(arr)
    ax_map.scatter(cx, cy, s=90, facecolor="white", edgecolor="black", linewidth=2.0, zorder=7)
    ax_map.scatter(cx, cy, s=20, facecolor="#0b2545", edgecolor="white", linewidth=0.5, zorder=8)
    ax_map.text(cx+0.12, cy+0.08, "Centroide", fontsize=7, fontweight="bold", color="#0b2545", zorder=9)

    ax_map.set_title(title, loc="left", fontsize=13, color=panel_color, fontweight="bold", pad=8)
    ax_map.set_xlim(1,5)
    ax_map.set_ylim(1,3)
    ax_map.set_xticks([1,2,3,4,5])
    ax_map.set_yticks([1,2,3])
    ax_map.set_yticklabels(["A\nInterior", "B\nCentro", "C\nExterior"], fontsize=7)
    ax_map.set_xlabel("Posición circunferencial (1 → 5)", fontsize=7)
    ax_map.tick_params(axis="x", labelsize=7)
    for s in ax_map.spines.values():
        s.set_edgecolor(panel_color)
        s.set_linewidth(1.0)

    # Table
    ax_tbl.axis("off")
    tbl_data = [["P","A","B","C"]] + [[str(i+1)] + [f"{arr[i,j]:.2f}" for j in range(3)] for i in range(5)]
    table = ax_tbl.table(cellText=tbl_data, loc="center", cellLoc="center")
    table.auto_set_font_size(False)
    table.set_fontsize(7)
    table.scale(1.0, 1.35)
    for (r,c), cell in table.get_celld().items():
        cell.set_linewidth(0.45)
        cell.set_edgecolor("#c5d0df")
        if r == 0:
            cell.set_facecolor("#0b2545")
            cell.get_text().set_color("white")
            cell.get_text().set_fontweight("bold")
    ax_tbl.set_title("LECTURA VISUAL\n(mm)", fontsize=8, color="#0b2545", fontweight="bold")

    # KPI
    ax_kpi.axis("off")
    mean = np.mean(arr)
    mn = np.min(arr)
    mx = np.max(arr)
    rg = mx - mn
    kpis = [("PROMEDIO", mean, "#0b2545"), ("MÍNIMO", mn, "red"), ("MÁXIMO", mx, "#168323"), ("RANGO", rg, "#0b2545")]
    y = 0.95
    for label, value, color in kpis:
        ax_kpi.text(0.5, y, label, ha="center", va="top", fontsize=7, color=color, fontweight="bold")
        ax_kpi.text(0.5, y-0.105, f"{value:.2f} mm", ha="center", va="top", fontsize=12, color=color, fontweight="bold")
        y -= 0.22

    cond, cond_color = condition(arr)
    ax_kpi.add_patch(plt.Rectangle((0.08, 0.02), 0.84, 0.16, transform=ax_kpi.transAxes, facecolor=cond_color, edgecolor=cond_color))
    ax_kpi.text(0.5, 0.12, "CONDICIÓN", ha="center", va="center", fontsize=7, color="white" if cond!="MONITOREAR" else "#0b2545", fontweight="bold")
    ax_kpi.text(0.5, 0.055, cond, ha="center", va="center", fontsize=10, color="white" if cond!="MONITOREAR" else "#0b2545", fontweight="bold")

def create_report(data, meta):
    cmap = LinearSegmentedColormap.from_list(
        "wear", ["#b30000", "#e34a33", "#fdbb84", "#fee08b", "#d9ef8b", "#66bd63", "#1a9850"]
    )

    vals = np.concatenate([v.flatten() for v in data.values()])
    vmin = np.floor((float(vals.min()) - 0.02)*100)/100
    vmax = np.ceil((float(vals.max()) + 0.02)*100)/100
    norm = Normalize(vmin=vmin, vmax=vmax)

    fig = plt.figure(figsize=(16,10), dpi=180)
    fig.patch.set_facecolor("white")

    gs = GridSpec(
        6, 2, figure=fig,
        height_ratios=[0.55, 0.45, 2.4, 2.4, 0.92, 0.18],
        width_ratios=[1,1],
        hspace=0.28,
        wspace=0.05
    )

    # Header
    ax_h = fig.add_subplot(gs[0,:])
    ax_h.axis("off")
    ax_h.text(0.0, 0.78, "LUFUSSA", fontsize=22, fontweight="bold", color="#5b6570", va="center")
    ax_h.text(0.20, 0.78, "INSPECCIÓN DE COJINETES BIG END BEARING – FIRMA SUPERFICIAL DE DESGASTE",
              fontsize=17, fontweight="bold", color="#0b2545", va="center")
    ax_h.text(0.27, 0.30, f"Cilindros {meta['cil_a']} y {meta['cil_b']}   |   Datos reales de medición   |   Rojo = menor espesor   /   Verde = mayor espesor",
              fontsize=11, color="#0b55b8", va="center")
    ax_h.text(0.90, 0.78, f"FECHA: {meta['fecha']}\nREPORTE: {meta['reporte']}\nINSPECTOR: {meta['inspector']}",
              fontsize=8, color="#0b2545", va="center", ha="left", fontweight="bold")

    # Equipment strip
    ax_e = fig.add_subplot(gs[1,:])
    ax_e.axis("off")
    items = [("EQUIPO / UNIDAD", meta["unidad"]), ("MOTOR", meta["motor"]), ("BANCO", "A y B"),
             ("HORAS DE OPERACIÓN", meta["horas"]), ("TIPO DE COJINETE", "BIG END BEARING")]
    x = 0.02
    widths = [0.16,0.18,0.12,0.20,0.22]
    for idx, (label, val) in enumerate(items):
        w = widths[idx]
        if idx == 0:
            ax_e.add_patch(plt.Rectangle((x, 0.10), w, 0.80, transform=ax_e.transAxes, facecolor="#0b2545", edgecolor="#0b2545"))
            col = "white"
        else:
            ax_e.add_patch(plt.Rectangle((x, 0.10), w, 0.80, transform=ax_e.transAxes, facecolor="white", edgecolor="#c9d4e4"))
            col = "#0b2545"
        ax_e.text(x+w/2, 0.62, label, ha="center", va="center", fontsize=7, fontweight="bold", color=col)
        ax_e.text(x+w/2, 0.32, val, ha="center", va="center", fontsize=11, fontweight="bold", color=col)
        x += w

    # Panels
    keys = list(data.keys())
    colors = []
    for k in keys:
        _, _, = condition(data[k])
    panel_colors = [condition(data[k])[1] for k in keys]
    add_panel(fig, gs[2,0], keys[0], data[keys[0]], cmap, norm, panel_colors[0])
    add_panel(fig, gs[2,1], keys[1], data[keys[1]], cmap, norm, panel_colors[1])
    add_panel(fig, gs[3,0], keys[2], data[keys[2]], cmap, norm, panel_colors[2])
    add_panel(fig, gs[3,1], keys[3], data[keys[3]], cmap, norm, panel_colors[3])

    # Footer
    gs_f = GridSpecFromSubplotSpec(1, 4, subplot_spec=gs[4,:], width_ratios=[1.0, 1.0, 1.35, 2.3], wspace=0.15)
    ax_scale = fig.add_subplot(gs_f[0,0]); ax_scale.axis("off")
    ax_sym = fig.add_subplot(gs_f[0,1]); ax_sym.axis("off")
    ax_risk = fig.add_subplot(gs_f[0,2]); ax_risk.axis("off")
    ax_diag = fig.add_subplot(gs_f[0,3]); ax_diag.axis("off")

    ax_scale.set_title("ESCALA DE ESPESOR", fontsize=9, color="#0b2545", fontweight="bold")
    grad = np.linspace(vmin, vmax, 256).reshape(-1,1)
    ax_in = ax_scale.inset_axes([0.12,0.12,0.10,0.72])
    ax_in.imshow(grad, cmap=cmap, norm=norm, origin="lower", aspect="auto")
    ax_in.set_xticks([])
    ax_in.set_yticks(np.linspace(0,255,5))
    ax_in.set_yticklabels([f"{v:.2f}" for v in np.linspace(vmin,vmax,5)], fontsize=7)
    for sp in ax_in.spines.values(): sp.set_visible(False)
    ax_scale.text(0.35,0.75,"MAYOR ESPESOR",fontsize=7,color="#168323",fontweight="bold",transform=ax_scale.transAxes)
    ax_scale.text(0.35,0.20,"MENOR ESPESOR",fontsize=7,color="red",fontweight="bold",transform=ax_scale.transAxes)

    ax_sym.set_title("SIMBOLOGÍA", fontsize=9, color="#0b2545", fontweight="bold")
    sym = [("○","Punto de medición"),("□","Menor espesor"),("◎","Centroide de desgaste"),("—","Curva de nivel")]
    for i,(a,b) in enumerate(sym):
        ax_sym.text(0.08,0.75-i*0.18,a,fontsize=11,color="black")
        ax_sym.text(0.22,0.76-i*0.18,b,fontsize=7.5,color="#0b2545")

    ax_risk.set_title("MAPA DE RIESGO (RESUMEN)", fontsize=9, color="#0b2545", fontweight="bold")
    for i,k in enumerate(keys):
        cond, col = condition(data[k])
        ax_risk.text(0.16+i*0.22,0.62,k.replace(" ","\n"),fontsize=7,color="#0b2545",ha="center",fontweight="bold")
        ax_risk.scatter(0.16+i*0.22,0.25,s=320,color=col,edgecolor="black")

    ax_diag.set_title("DIAGNÓSTICO RESUMEN", fontsize=9, color="#0b2545", fontweight="bold")
    diag_lines = []
    for k in keys:
        cond, col = condition(data[k])
        if cond == "NORMAL":
            txt = f"{k}: Desgaste uniforme. Sin anomalías relevantes."
        elif cond == "MONITOREAR":
            txt = f"{k}: Condición moderada. Monitorear tendencia en próximos ciclos."
        else:
            txt = f"{k}: Desgaste localizado severo. Recomendada inspección detallada."
        diag_lines.append((col, txt))
    for i,(col,txt) in enumerate(diag_lines[:3]):
        ax_diag.scatter(0.03,0.78-i*0.25,s=220,color=col,edgecolor="black")
        ax_diag.text(0.08,0.78-i*0.25,txt,fontsize=8,color="#0b2545",va="center")

    # Note
    ax_n = fig.add_subplot(gs[5,:])
    ax_n.axis("off")
    ax_n.text(0.0,0.5,"NOTA: Medidas en mm. Valores interpolados como apoyo visual; los puntos de medición representan los datos reales.",
              fontsize=7.5,color="#0b2545")
    ax_n.text(0.88,0.5,"LUFUSSA  |  CONFIDENCIAL",fontsize=8,color="#65758b")

    return fig

# ==================================================
# INTERFAZ
# ==================================================
st.markdown("""
<div class="header-box">
<div class="title">OEM Insight Studio Web v1.2</div>
<div class="subtitle">Generador de reporte visual Big End Bearing</div>
</div>
""", unsafe_allow_html=True)

st.info("Versión 1.2: corrige el problema de mapas en blanco y ordena el reporte con una estructura más estable.")

c1, c2 = st.columns(2)
with c1:
    file_a = st.file_uploader("Excel Banco A", type=["xlsx"], key="a")
with c2:
    file_b = st.file_uploader("Excel Banco B", type=["xlsx"], key="b")

st.subheader("Datos del reporte")
m1, m2, m3 = st.columns(3)
with m1:
    unidad = st.text_input("Equipo / Unidad", "GENSET 11")
    fecha = st.text_input("Fecha", date.today().strftime("%d / %m / %Y"))
with m2:
    motor = st.text_input("Motor", "WÄRTSILÄ 46")
    reporte = st.text_input("Reporte No.", "WSM-2026-0702")
with m3:
    horas = st.text_input("Horas de operación", "12,845 h")
    inspector = st.text_input("Inspector", "")

if file_a and file_b:
    sheets_a = get_sheets(file_a)
    sheets_b = get_sheets(file_b)
    s1, s2 = st.columns(2)
    with s1:
        sheet_a = st.selectbox("Hoja Banco A", sheets_a)
    with s2:
        sheet_b = st.selectbox("Hoja Banco B", sheets_b)

    mode = st.radio("Modo de lectura", ["Plantilla original", "Manual"], horizontal=True)

    if mode == "Manual":
        st.warning("Usa el número de fila y columna de la esquina superior izquierda de cada matriz 5x3.")
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            au_r = st.number_input("A Upper fila", value=14, min_value=1)
            au_c = st.number_input("A Upper columna", value=7, min_value=1)
        with col2:
            al_r = st.number_input("A Lower fila", value=14, min_value=1)
            al_c = st.number_input("A Lower columna", value=18, min_value=1)
        with col3:
            bu_r = st.number_input("B Upper fila", value=14, min_value=1)
            bu_c = st.number_input("B Upper columna", value=7, min_value=1)
        with col4:
            bl_r = st.number_input("B Lower fila", value=14, min_value=1)
            bl_c = st.number_input("B Lower columna", value=18, min_value=1)

    cil_a = st.text_input("Nombre cilindro Banco A", "A1")
    cil_b = st.text_input("Nombre cilindro Banco B", "B1")

    if st.button("Generar reporte", type="primary"):
        try:
            if mode == "Plantilla original":
                arr_au = read_template(file_a, sheet_a, True)
                arr_al = read_template(file_a, sheet_a, False)
                arr_bu = read_template(file_b, sheet_b, True)
                arr_bl = read_template(file_b, sheet_b, False)
            else:
                arr_au = read_manual(file_a, sheet_a, int(au_r), int(au_c))
                arr_al = read_manual(file_a, sheet_a, int(al_r), int(al_c))
                arr_bu = read_manual(file_b, sheet_b, int(bu_r), int(bu_c))
                arr_bl = read_manual(file_b, sheet_b, int(bl_r), int(bl_c))

            data = {
                f"{cil_a} UPPER": arr_au,
                f"{cil_a} LOWER": arr_al,
                f"{cil_b} UPPER": arr_bu,
                f"{cil_b} LOWER": arr_bl
            }
            meta = {
                "unidad": unidad, "motor": motor, "horas": horas,
                "fecha": fecha, "reporte": reporte, "inspector": inspector,
                "cil_a": cil_a, "cil_b": cil_b
            }

            fig = create_report(data, meta)
            png = BytesIO()
            pdf = BytesIO()
            fig.savefig(png, format="png", bbox_inches="tight", facecolor="white")
            fig.savefig(pdf, format="pdf", bbox_inches="tight", facecolor="white")
            png.seek(0); pdf.seek(0)
            plt.close(fig)

            st.success("Reporte generado correctamente.")
            st.image(png, use_container_width=True)
            st.download_button("Descargar PNG", data=png.getvalue(), file_name="Reporte_BEB_WSM.png", mime="image/png")
            st.download_button("Descargar PDF", data=pdf.getvalue(), file_name="Reporte_BEB_WSM.pdf", mime="application/pdf")

        except Exception as e:
            st.error(f"No se pudo generar el reporte: {e}")
else:
    st.warning("Carga ambos archivos Excel para continuar.")
