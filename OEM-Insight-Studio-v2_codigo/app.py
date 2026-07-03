
import streamlit as st
import numpy as np
import openpyxl
import matplotlib.pyplot as plt
from matplotlib.colors import LinearSegmentedColormap, Normalize
from matplotlib.patches import FancyBboxPatch, Circle
from io import BytesIO
from datetime import date
from pathlib import Path

st.set_page_config(
    page_title="OEM Insight Studio",
    page_icon="📊",
    layout="wide"
)

st.markdown("""
<style>
.stApp {
    background: #f3f6fb;
}
.main-card {
    background: white;
    padding: 22px;
    border-radius: 18px;
    border: 1px solid #d9e2ef;
    box-shadow: 0 4px 14px rgba(15,23,42,.06);
}
.title {
    color: #0b2545;
    font-size: 34px;
    font-weight: 800;
}
.subtitle {
    color: #31517a;
    font-size: 16px;
}
</style>
""", unsafe_allow_html=True)

# ==========================================================
# LECTURA DE EXCEL
# ==========================================================
def load_workbook(uploaded_file):
    uploaded_file.seek(0)
    return openpyxl.load_workbook(uploaded_file, data_only=True)

def get_sheet_names(uploaded_file):
    return load_workbook(uploaded_file).sheetnames

def read_matrix_template(uploaded_file, sheet_name, upper=True):
    """
    Lee matriz 5x3 desde la estructura original del Excel:
    UPPER: columnas G:I, filas 14:18
    LOWER: columnas R:T, filas 14:18
    """
    wb = load_workbook(uploaded_file)
    ws = wb[sheet_name]

    rows = range(14, 19)
    cols = [7, 8, 9] if upper else [18, 19, 20]

    values = []
    for r in rows:
        row = []
        for c in cols:
            val = ws.cell(r, c).value
            if val is None:
                raise ValueError(f"Celda vacía en hoja {sheet_name}, fila {r}, columna {c}")
            row.append(float(val))
        values.append(row)

    return np.array(values, dtype=float)

def read_matrix_manual(uploaded_file, sheet_name, start_row, start_col):
    wb = load_workbook(uploaded_file)
    ws = wb[sheet_name]

    values = []
    for r in range(start_row, start_row + 5):
        row = []
        for c in range(start_col, start_col + 3):
            row.append(float(ws.cell(r, c).value))
        values.append(row)

    return np.array(values, dtype=float)

# ==========================================================
# CÁLCULOS
# ==========================================================
def condition(arr):
    mn = float(np.min(arr))
    rg = float(np.max(arr) - np.min(arr))

    if mn < 0.85 or rg > 0.50:
        return "REVISAR", "#e10600", "#d7191c"
    elif mn < 1.05 or rg > 0.28:
        return "MONITOREAR", "#f5b400", "#b7791f"
    else:
        return "NORMAL", "#228b22", "#188038"

def metrics(arr):
    return {
        "mean": float(np.mean(arr)),
        "min": float(np.min(arr)),
        "max": float(np.max(arr)),
        "range": float(np.max(arr) - np.min(arr))
    }

def interp_surface(arr, nx=220, ny=150):
    x = np.arange(1, 6)
    y = np.arange(1, 4)

    xi = np.linspace(1, 5, nx)
    temp = np.array([np.interp(xi, x, arr[:, j]) for j in range(3)])

    yi = np.linspace(1, 3, ny)
    zi = np.array([np.interp(yi, y, temp[:, i]) for i in range(nx)]).T

    X, Y = np.meshgrid(xi, yi)
    return X, Y, zi

def wear_centroid(arr):
    wear = arr.max() - arr
    total = wear.sum()

    Xp, Yp = np.meshgrid(
        np.arange(1, 6),
        np.arange(1, 4),
        indexing="ij"
    )

    if total <= 0:
        return 3, 2

    cx = (Xp * wear).sum() / total
    cy = (Yp * wear).sum() / total
    return cx, cy

# ==========================================================
# REPORTE
# ==========================================================
def create_report(data, meta):
    cmap = LinearSegmentedColormap.from_list(
        "wear_scale",
        [
            "#b30000",
            "#e34a33",
            "#fdbb84",
            "#fee08b",
            "#d9ef8b",
            "#66bd63",
            "#1a9850"
        ]
    )

    all_values = np.concatenate([v.flatten() for v in data.values()])
    vmin = np.floor((float(all_values.min()) - 0.02) * 100) / 100
    vmax = np.ceil((float(all_values.max()) + 0.02) * 100) / 100
    norm = Normalize(vmin=vmin, vmax=vmax)

    fig = plt.figure(figsize=(16, 10), dpi=180)
    fig.patch.set_facecolor("white")

    # -------------------------
    # Encabezado
    # -------------------------
    logo_path = Path("assets/lufussa_logo.png")
    ax_logo = fig.add_axes([0.025, 0.905, 0.145, 0.080])
    ax_logo.axis("off")
    if logo_path.exists():
        logo = plt.imread(str(logo_path))
        ax_logo.imshow(logo)
    else:
        ax_logo.text(0, 0.5, "LUFUSSA", fontsize=24, fontweight="bold", color="#6b6f76")

    fig.text(
        0.260,
        0.955,
        "MEDICIÓN DE COJINETES BIG END BERING",
        fontsize=22,
        fontweight="bold",
        color="#0b2545",
        ha="left",
        va="center"
    )

    fig.text(
        0.250,
        0.925,
        f"Cilindros {meta['cil_a']} y {meta['cil_b']}   |   Datos reales de medición   |   Rojo = menor espesor   |   Verde = mayor espesor",
        fontsize=11.2,
        color="#0b55b8",
        ha="left",
        va="center"
    )

    ax_meta = fig.add_axes([0.835, 0.885, 0.145, 0.110])
    ax_meta.axis("off")
    ax_meta.add_patch(
        FancyBboxPatch(
            (0, 0),
            1,
            1,
            boxstyle="round,pad=0.015",
            facecolor="#ffffff",
            edgecolor="#c9d4e4",
            linewidth=0.8
        )
    )
    ax_meta.text(0.08, 0.72, "FECHA:", fontsize=8, color="#0b2545", fontweight="bold")
    ax_meta.text(0.55, 0.72, meta["fecha"], fontsize=8, color="#0b2545", fontweight="bold", ha="center")
    ax_meta.text(0.08, 0.45, "REPORTE:", fontsize=8, color="#0b2545", fontweight="bold")
    ax_meta.text(0.58, 0.45, meta["reporte"], fontsize=8, color="#0b2545", fontweight="bold", ha="center")
    ax_meta.text(0.08, 0.18, "INSPECTOR:", fontsize=8, color="#0b2545", fontweight="bold")
    ax_meta.text(0.58, 0.18, meta["inspector"] if meta["inspector"] else "__________", fontsize=8, color="#0b2545", ha="center")

    # -------------------------
    # Banda de equipo
    # -------------------------
    ax_eq = fig.add_axes([0.025, 0.815, 0.950, 0.060])
    ax_eq.axis("off")
    ax_eq.add_patch(
        FancyBboxPatch(
            (0, 0),
            1,
            1,
            boxstyle="round,pad=0.008",
            facecolor="white",
            edgecolor="#c9d4e4",
            linewidth=0.8
        )
    )

    fields = [
        ("EQUIPO / UNIDAD", meta["unidad"], 0.18, True),
        ("MOTOR", meta["motor"], 0.22, False),
        ("BANCO", "A y B", 0.16, False),
        ("HORAS DE OPERACIÓN", meta["horas"], 0.22, False),
        ("TIPO DE COJINETE", "BIG END BEARING", 0.22, False)
    ]

    x = 0.010
    for label, value, width, dark in fields:
        if dark:
            ax_eq.add_patch(
                FancyBboxPatch(
                    (x, 0.12),
                    width,
                    0.76,
                    boxstyle="round,pad=0.006",
                    facecolor="#0b2545",
                    edgecolor="#0b2545"
                )
            )
            color = "white"
        else:
            ax_eq.add_line(
                plt.Line2D([x, x], [0.15, 0.85], color="#c9d4e4", linewidth=1)
            )
            color = "#0b2545"

        ax_eq.text(x + width / 2, 0.62, label, ha="center", va="center", fontsize=7.2, color=color, fontweight="bold")
        ax_eq.text(x + width / 2, 0.34, value, ha="center", va="center", fontsize=11, color=color, fontweight="bold")
        x += width

    # -------------------------
    # Panel individual
    # -------------------------
    def draw_panel(title, arr, x, y, w, h):
        cond_label, cond_color, border_color = condition(arr)
        k = metrics(arr)

        ax_box = fig.add_axes([x, y, w, h])
        ax_box.axis("off")
        ax_box.add_patch(
            FancyBboxPatch(
                (0, 0),
                1,
                1,
                boxstyle="round,pad=0.006",
                facecolor="white",
                edgecolor=border_color,
                linewidth=1.0
            )
        )
        ax_box.text(0.045, 0.94, title, fontsize=13.0, color=border_color, fontweight="bold", va="top")

        # Mapa
        ax_map = fig.add_axes([x + w * 0.055, y + h * 0.170, w * 0.455, h * 0.670])
        X, Y, Z = interp_surface(arr)

        ax_map.imshow(
            Z.T,
            origin="lower",
            extent=[1, 5, 1, 3],
            cmap=cmap,
            norm=norm,
            interpolation="bicubic",
            aspect="auto"
        )

        levels = np.linspace(vmin, vmax, 5)
        cs = ax_map.contour(X, Y, Z, levels=levels, colors="#0b2545", linewidths=0.42, alpha=0.32)
        ax_map.clabel(cs, fontsize=6, inline=True, fmt="%.2f")

        # puntos medidos
        for i in range(5):
            for j in range(3):
                ax_map.scatter(i + 1, j + 1, s=22, facecolor="white", edgecolor="black", linewidth=0.75, zorder=4)

        # punto mínimo
        mi = np.unravel_index(np.argmin(arr), arr.shape)
        ax_map.scatter(mi[0] + 1, mi[1] + 1, s=70, marker="s", facecolor="white", edgecolor="black", linewidth=1.2, zorder=5)

        # centroide
        cx, cy = wear_centroid(arr)
        ax_map.scatter(cx, cy, s=80, facecolor="white", edgecolor="black", linewidth=2.0, zorder=6)
        ax_map.scatter(cx, cy, s=18, facecolor="#0b2545", edgecolor="white", linewidth=0.5, zorder=7)
        ax_map.text(cx + 0.12, cy + 0.08, "Centroide", fontsize=6.5, fontweight="bold", color="#0b2545", zorder=8)

        ax_map.set_xlim(1, 5)
        ax_map.set_ylim(1, 3)
        ax_map.set_xticks([1, 2, 3, 4, 5])
        ax_map.set_yticks([1, 2, 3])
        ax_map.set_yticklabels(["A\nInterior", "B\nCentro", "C\nExterior"], fontsize=7)
        ax_map.set_xlabel("Posición circunferencial (1 → 5)", fontsize=7)
        ax_map.tick_params(axis="x", labelsize=7)
        for sp in ax_map.spines.values():
            sp.set_visible(False)

        # Tabla
        ax_tbl = fig.add_axes([x + w * 0.565, y + h * 0.280, w * 0.205, h * 0.405])
        ax_tbl.axis("off")

        table_data = [["P", "A", "B", "C"]]
        for i in range(5):
            table_data.append([str(i + 1)] + [f"{arr[i, j]:.2f}" for j in range(3)])

        table = ax_tbl.table(cellText=table_data, loc="center", cellLoc="center")
        table.auto_set_font_size(False)
        table.set_fontsize(6.8)
        table.scale(1.00, 1.20)

        for (r, c), cell in table.get_celld().items():
            cell.set_edgecolor("#c5d0df")
            cell.set_linewidth(0.45)
            if r == 0:
                cell.set_facecolor("#0b2545")
                cell.get_text().set_color("white")
                cell.get_text().set_fontweight("bold")

        ax_tbl.set_title("LECTURA VISUAL (mm)", fontsize=7.5, color="#0b2545", fontweight="bold", pad=3)

        # KPIs
        kpi_x = x + w * 0.805
        kpi_y = y + h * 0.250
        kpi_w = w * 0.145
        kpi_h = h * 0.565

        kpis = [
            ("PROMEDIO", k["mean"], "#0b2545"),
            ("MÍNIMO", k["min"], "red"),
            ("MÁXIMO", k["max"], "#168323"),
            ("RANGO", k["range"], "#0b2545")
        ]

        for idx, (label, value, color) in enumerate(kpis):
            yy = kpi_y + kpi_h - (idx + 1) * h * 0.145
            fig.add_artist(
                FancyBboxPatch(
                    (kpi_x, yy),
                    kpi_w,
                    h * 0.120,
                    transform=fig.transFigure,
                    boxstyle="round,pad=0.003,rounding_size=0.004",
                    facecolor="#f8fafc",
                    edgecolor="#d7e0ec",
                    linewidth=0.55
                )
            )
            fig.text(kpi_x + kpi_w / 2, yy + h * 0.078, label, fontsize=5.8, color=color, ha="center", fontweight="bold")
            fig.text(kpi_x + kpi_w / 2, yy + h * 0.028, f"{value:.2f}", fontsize=10, color=color, ha="center", fontweight="bold")
            fig.text(kpi_x + kpi_w * 0.78, yy + h * 0.032, "mm", fontsize=5.2, color=color, ha="left", fontweight="bold")

        # Condición
        cond_x = x + w * 0.735
        cond_y = y + h * 0.045
        cond_w = w * 0.225
        cond_h = h * 0.125

        fig.add_artist(
            FancyBboxPatch(
                (cond_x, cond_y),
                cond_w,
                cond_h,
                transform=fig.transFigure,
                boxstyle="round,pad=0.004,rounding_size=0.006",
                facecolor=cond_color,
                edgecolor=cond_color,
                linewidth=0.7
            )
        )

        fig.add_artist(
            Circle(
                (cond_x + cond_w * 0.18, cond_y + cond_h * 0.50),
                cond_h * 0.23,
                transform=fig.transFigure,
                facecolor="none",
                edgecolor="white" if cond_label != "MONITOREAR" else "black",
                linewidth=1.0
            )
        )

        txt_color = "white" if cond_label != "MONITOREAR" else "#0b2545"

        fig.text(cond_x + cond_w * 0.62, cond_y + cond_h * 0.65, "CONDICIÓN", fontsize=6.2, color=txt_color, ha="center", fontweight="bold")
        fig.text(cond_x + cond_w * 0.62, cond_y + cond_h * 0.31, cond_label, fontsize=9, color=txt_color, ha="center", fontweight="bold")

    keys = list(data.keys())
    draw_panel(keys[0], data[keys[0]], 0.025, 0.515, 0.465, 0.285)
    draw_panel(keys[1], data[keys[1]], 0.510, 0.515, 0.465, 0.285)
    draw_panel(keys[2], data[keys[2]], 0.025, 0.215, 0.465, 0.285)
    draw_panel(keys[3], data[keys[3]], 0.510, 0.215, 0.465, 0.285)

    # -------------------------
    # Pie de reporte sin nota ni confidencial
    # -------------------------
    footer_y = 0.055
    footer_h = 0.135

    # Escala
    ax_scale = fig.add_axes([0.025, footer_y, 0.290, footer_h])
    ax_scale.axis("off")
    ax_scale.add_patch(FancyBboxPatch((0, 0), 1, 1, boxstyle="round,pad=0.006", facecolor="white", edgecolor="#d7e0ec", linewidth=0.8))
    ax_scale.text(0.08, 0.87, "ESCALA DE ESPESOR", fontsize=8.5, color="#0b2545", fontweight="bold")

    cax = ax_scale.inset_axes([0.07, 0.16, 0.055, 0.66])
    grad = np.linspace(vmin, vmax, 256).reshape(-1, 1)
    cax.imshow(grad, cmap=cmap, norm=norm, origin="lower", aspect="auto")
    cax.set_xticks([])
    cax.set_yticks([])
    for sp in cax.spines.values():
        sp.set_visible(False)

    ax_scale.text(0.30, 0.70, "MAYOR ESPESOR", fontsize=7, color="#168323", fontweight="bold")
    ax_scale.text(0.30, 0.22, "MENOR ESPESOR", fontsize=7, color="red", fontweight="bold")

    # Simbología
    ax_sym = fig.add_axes([0.325, footer_y, 0.185, footer_h])
    ax_sym.axis("off")
    ax_sym.add_patch(FancyBboxPatch((0, 0), 1, 1, boxstyle="round,pad=0.006", facecolor="white", edgecolor="#d7e0ec", linewidth=0.8))
    ax_sym.text(0.12, 0.87, "SIMBOLOGÍA", fontsize=8.5, color="#0b2545", fontweight="bold")

    symbols = [
        ("○", "Punto de medición"),
        ("□", "Menor espesor"),
        ("◎", "Centroide de desgaste"),
        ("—", "Curva de nivel")
    ]

    for i, (symbol, text) in enumerate(symbols):
        yy = 0.67 - i * 0.17
        ax_sym.text(0.12, yy, symbol, fontsize=10, color="black")
        ax_sym.text(0.26, yy + 0.01, text, fontsize=7, color="#0b2545")

    # Mapa de riesgo
    ax_risk = fig.add_axes([0.520, footer_y, 0.220, footer_h])
    ax_risk.axis("off")
    ax_risk.add_patch(FancyBboxPatch((0, 0), 1, 1, boxstyle="round,pad=0.006", facecolor="white", edgecolor="#d7e0ec", linewidth=0.8))
    ax_risk.text(0.26, 0.87, "MAPA DE RIESGO (RESUMEN)", fontsize=8.5, color="#0b2545", fontweight="bold")

    for i, key in enumerate(keys):
        cond_label, cond_color, _ = condition(data[key])
        xx = 0.13 + i * 0.22
        ax_risk.text(xx, 0.58, key.replace(" ", "\n"), fontsize=6.8, color="#0b2545", ha="center", fontweight="bold")
        ax_risk.scatter(xx, 0.25, s=300, color=cond_color, edgecolor="black")

    # Diagnóstico
    ax_diag = fig.add_axes([0.750, footer_y, 0.225, footer_h])
    ax_diag.axis("off")
    ax_diag.add_patch(FancyBboxPatch((0, 0), 1, 1, boxstyle="round,pad=0.006", facecolor="white", edgecolor="#d7e0ec", linewidth=0.8))
    ax_diag.text(0.20, 0.87, "DIAGNÓSTICO RESUMEN", fontsize=8.5, color="#0b2545", fontweight="bold")

    for i, key in enumerate(keys):
        cond_label, cond_color, _ = condition(data[key])

        if cond_label == "NORMAL":
            text = f"{key}: desgaste uniforme."
        elif cond_label == "MONITOREAR":
            text = f"{key}: monitorear tendencia."
        else:
            text = f"{key}: revisar condición."

        yy = 0.66 - i * 0.18
        ax_diag.scatter(0.08, yy, s=120, color=cond_color, edgecolor="black")
        ax_diag.text(0.15, yy, text, fontsize=6.8, color="#0b2545", va="center")

    return fig

# ==========================================================
# INTERFAZ STREAMLIT
# ==========================================================
st.markdown("""
<div class="main-card">
<div class="title">OEM Insight Studio</div>
<div class="subtitle">Generador de reporte visual para medición de cojinetes Big End Bearing</div>
</div>
""", unsafe_allow_html=True)

st.write("")

c1, c2 = st.columns(2)
with c1:
    file_a = st.file_uploader("Excel Banco A", type=["xlsx"], key="a")
with c2:
    file_b = st.file_uploader("Excel Banco B", type=["xlsx"], key="b")

st.subheader("Datos del reporte")

m1, m2, m3 = st.columns(3)
with m1:
    unidad = st.text_input("Equipo / Unidad", "GENSET 11")
    fecha = st.text_input("Fecha", date.today().strftime("%d / %m / %Y"))
with m2:
    motor = st.text_input("Motor", "WÄRTSILÄ 46")
    reporte = st.text_input("Reporte No.", "MEB-2026-0702")
with m3:
    horas = st.text_input("Horas de operación", "12,845 h")
    inspector = st.text_input("Inspector", "")

if file_a and file_b:
    sheets_a = get_sheet_names(file_a)
    sheets_b = get_sheet_names(file_b)

    s1, s2 = st.columns(2)
    with s1:
        sheet_a = st.selectbox("Hoja Banco A", sheets_a)
    with s2:
        sheet_b = st.selectbox("Hoja Banco B", sheets_b)

    mode = st.radio("Modo de lectura", ["Plantilla original", "Manual"], horizontal=True)

    if mode == "Manual":
        st.warning("Indica fila y columna de la esquina superior izquierda de cada matriz 5x3.")
        col1, col2, col3, col4 = st.columns(4)

        with col1:
            au_r = st.number_input("A Upper fila", value=14, min_value=1)
            au_c = st.number_input("A Upper columna", value=7, min_value=1)

        with col2:
            al_r = st.number_input("A Lower fila", value=14, min_value=1)
            al_c = st.number_input("A Lower columna", value=18, min_value=1)

        with col3:
            bu_r = st.number_input("B Upper fila", value=14, min_value=1)
            bu_c = st.number_input("B Upper columna", value=7, min_value=1)

        with col4:
            bl_r = st.number_input("B Lower fila", value=14, min_value=1)
            bl_c = st.number_input("B Lower columna", value=18, min_value=1)

    cil_a = st.text_input("Nombre cilindro Banco A", "A1")
    cil_b = st.text_input("Nombre cilindro Banco B", "B1")

    if st.button("Generar reporte", type="primary"):
        try:
            if mode == "Plantilla original":
                arr_au = read_matrix_template(file_a, sheet_a, True)
                arr_al = read_matrix_template(file_a, sheet_a, False)
                arr_bu = read_matrix_template(file_b, sheet_b, True)
                arr_bl = read_matrix_template(file_b, sheet_b, False)
            else:
                arr_au = read_matrix_manual(file_a, sheet_a, int(au_r), int(au_c))
                arr_al = read_matrix_manual(file_a, sheet_a, int(al_r), int(al_c))
                arr_bu = read_matrix_manual(file_b, sheet_b, int(bu_r), int(bu_c))
                arr_bl = read_matrix_manual(file_b, sheet_b, int(bl_r), int(bl_c))

            data = {
                f"{cil_a} UPPER": arr_au,
                f"{cil_a} LOWER": arr_al,
                f"{cil_b} UPPER": arr_bu,
                f"{cil_b} LOWER": arr_bl
            }

            meta = {
                "unidad": unidad,
                "motor": motor,
                "horas": horas,
                "fecha": fecha,
                "reporte": reporte,
                "inspector": inspector,
                "cil_a": cil_a,
                "cil_b": cil_b
            }

            fig = create_report(data, meta)

            png = BytesIO()
            pdf = BytesIO()

            fig.savefig(png, format="png", bbox_inches="tight", facecolor="white")
            fig.savefig(pdf, format="pdf", bbox_inches="tight", facecolor="white")

            png.seek(0)
            pdf.seek(0)

            plt.close(fig)

            st.success("Reporte generado correctamente.")
            st.image(png, use_container_width=True)

            st.download_button(
                "Descargar PNG",
                data=png.getvalue(),
                file_name="Reporte_Medicion_Cojinetes_BEB.png",
                mime="image/png"
            )

            st.download_button(
                "Descargar PDF",
                data=pdf.getvalue(),
                file_name="Reporte_Medicion_Cojinetes_BEB.pdf",
                mime="application/pdf"
            )

        except Exception as e:
            st.error(f"No se pudo generar el reporte: {e}")
else:
    st.info("Carga ambos archivos Excel para comenzar.")
