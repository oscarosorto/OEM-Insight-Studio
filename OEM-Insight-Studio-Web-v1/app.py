
import streamlit as st
import pandas as pd
import numpy as np
import openpyxl
import matplotlib.pyplot as plt
from matplotlib.colors import LinearSegmentedColormap, Normalize
from matplotlib.patches import FancyBboxPatch, Circle, Rectangle
from io import BytesIO
from pathlib import Path
from datetime import date

st.set_page_config(
    page_title="OEM Insight Studio - Big End Bearing",
    page_icon="📊",
    layout="wide"
)

# =========================
# ESTILO GENERAL
# =========================
st.markdown("""
<style>
    .stApp {
        background: #f3f6fb;
    }
    .main-title {
        background: white;
        border-radius: 18px;
        padding: 18px 22px;
        border: 1px solid #d9e2ef;
        box-shadow: 0 4px 14px rgba(15, 23, 42, 0.06);
        margin-bottom: 20px;
    }
    .title-text {
        color: #0b2545;
        font-size: 32px;
        font-weight: 800;
        margin-bottom: 0px;
    }
    .subtitle-text {
        color: #31517a;
        font-size: 16px;
        margin-top: 4px;
    }
    .small-note {
        color: #64748b;
        font-size: 13px;
    }
</style>
""", unsafe_allow_html=True)

# =========================
# UTILIDADES
# =========================
def read_bearing(uploaded_file, sheet_name: str, upper: bool = True):
    wb = openpyxl.load_workbook(uploaded_file, data_only=True)
    ws = wb[sheet_name]
    rows = range(14, 19)
    cols = [7, 8, 9] if upper else [18, 19, 20]
    values = []
    for r in rows:
        row = []
        for c in cols:
            v = ws.cell(r, c).value
            if v is None:
                raise ValueError(f"Celda vacía en {sheet_name}: fila {r}, columna {c}")
            row.append(float(v))
        values.append(row)
    return np.array(values)

def find_sheet(workbook_file, cylinder_code: str):
    wb = openpyxl.load_workbook(workbook_file, data_only=True)
    candidates = wb.sheetnames
    for s in candidates:
        if cylinder_code.upper() in s.upper():
            return s
    raise ValueError(f"No se encontró hoja para {cylinder_code}. Hojas disponibles: {candidates}")

def interp_surface(arr, nx=220, ny=150):
    x = np.arange(1, 6)
    y = np.arange(1, 4)
    xi = np.linspace(1, 5, nx)
    temp = np.array([np.interp(xi, x, arr[:, j]) for j in range(3)])
    yi = np.linspace(1, 3, ny)
    zi = np.array([np.interp(yi, y, temp[:, i]) for i in range(nx)]).T
    X, Y = np.meshgrid(xi, yi)
    return X, Y, zi

def get_metrics(arr):
    return {
        "mean": float(np.mean(arr)),
        "min": float(np.min(arr)),
        "max": float(np.max(arr)),
        "range": float(np.max(arr) - np.min(arr)),
    }

def condition_from_values(arr):
    mn = float(np.min(arr))
    rg = float(np.max(arr) - np.min(arr))
    if mn < 0.85 or rg > 0.50:
        return "REVISAR", "#e10600"
    elif mn < 1.05 or rg > 0.28:
        return "MONITOREAR", "#f5b400"
    else:
        return "NORMAL", "#228b22"

def build_report(data, meta):
    # Colormap: rojo = menor espesor / mayor desgaste; verde = mayor espesor
    cmap = LinearSegmentedColormap.from_list(
        "wear",
        ["#b30000", "#e34a33", "#fdbb84", "#fee08b", "#d9ef8b", "#66bd63", "#1a9850"]
    )

    all_values = np.concatenate([v.flatten() for v in data.values()])
    vmin = np.floor((float(all_values.min()) - 0.02) * 100) / 100
    vmax = np.ceil((float(all_values.max()) + 0.02) * 100) / 100
    norm = Normalize(vmin=vmin, vmax=vmax)

    fig = plt.figure(figsize=(16, 10), dpi=180)
    fig.patch.set_facecolor("white")

    # Header
    fig.text(0.19, 0.955,
             "INSPECCIÓN DE COJINETES BIG END BEARING – FIRMA SUPERFICIAL DE DESGASTE",
             fontsize=18, fontweight="bold", color="#0b2545", ha="left")
    fig.text(0.25, 0.925,
             f"Cilindros {meta['cil_a']} y {meta['cil_b']}   |   Datos reales de medición   |   Rojo = menor espesor   /   Verde = mayor espesor",
             fontsize=11.5, color="#0b55b8", ha="left")

    # Logo fallback / text
    logo_file = Path("assets/lufussa_logo.png")
    if logo_file.exists():
        try:
            logo = plt.imread(str(logo_file))
            ax_logo = fig.add_axes([0.025, 0.91, 0.14, 0.075])
            ax_logo.imshow(logo)
            ax_logo.axis("off")
        except Exception:
            fig.text(0.025, 0.945, "LUFUSSA", fontsize=22, fontweight="bold", color="#5b6570")
    else:
        fig.text(0.025, 0.945, "LUFUSSA", fontsize=22, fontweight="bold", color="#5b6570")

    # Right box
    fig.add_artist(FancyBboxPatch((0.875, 0.895), 0.105, 0.09, transform=fig.transFigure,
                                  boxstyle="round,pad=0.006,rounding_size=0.006",
                                  facecolor="#ffffff", edgecolor="#c9d4e4", linewidth=0.8))
    fig.text(0.888, 0.963, "FECHA:", fontsize=7.5, color="#0b2545", fontweight="bold")
    fig.text(0.925, 0.963, meta["fecha"], fontsize=7.5, color="#0b2545", fontweight="bold")
    fig.text(0.888, 0.935, "REPORTE No.:", fontsize=7.5, color="#0b2545", fontweight="bold")
    fig.text(0.925, 0.935, meta["reporte"], fontsize=7.5, color="#0b2545", fontweight="bold")
    fig.text(0.888, 0.907, "INSPECTOR:", fontsize=7.5, color="#0b2545", fontweight="bold")
    fig.text(0.925, 0.907, meta["inspector"], fontsize=7.5, color="#0b2545")

    # Equipment strip
    strip_y = 0.83
    fig.add_artist(FancyBboxPatch((0.025, strip_y), 0.83, 0.055, transform=fig.transFigure,
                                  boxstyle="round,pad=0.004,rounding_size=0.006",
                                  facecolor="#ffffff", edgecolor="#c9d4e4", linewidth=0.8))
    items = [
        ("EQUIPO / UNIDAD", meta["unidad"], "#0b2545", "white"),
        ("MOTOR", meta["motor"], None, None),
        ("BANCO", "A y B", None, None),
        ("HORAS DE OPERACIÓN", meta["horas"], None, None),
        ("TIPO DE COJINETE", "BIG END BEARING", None, None),
    ]
    x0 = 0.035
    widths = [0.13, 0.16, 0.14, 0.18, 0.18]
    for i, (label, value, bg, fg) in enumerate(items):
        x = x0 + sum(widths[:i])
        w = widths[i]
        if bg:
            fig.add_artist(FancyBboxPatch((x, strip_y+0.005), w, 0.045, transform=fig.transFigure,
                                          boxstyle="round,pad=0.003,rounding_size=0.004",
                                          facecolor=bg, edgecolor=bg))
            lc = vc = fg
        else:
            lc = vc = "#0b2545"
            fig.add_artist(Rectangle((x, strip_y+0.005), 0.001, 0.045,
                                     transform=fig.transFigure, facecolor="#c9d4e4", edgecolor="none"))
        fig.text(x+w/2, strip_y+0.035, label, fontsize=7.5, color=lc, ha="center", fontweight="bold")
        fig.text(x+w/2, strip_y+0.013, value, fontsize=11, color=vc, ha="center", fontweight="bold")

    # Panel helper
    def panel(title, arr, pos, border_color):
        x, y, w, h = pos
        fig.add_artist(FancyBboxPatch((x, y), w, h, transform=fig.transFigure,
                                      boxstyle="round,pad=0.006,rounding_size=0.008",
                                      facecolor="#ffffff", edgecolor=border_color, linewidth=1.0))
        fig.text(x+0.025, y+h-0.028, title, fontsize=13, color=border_color, fontweight="bold")

        # map
        ax = fig.add_axes([x+0.035, y+0.055, w*0.46, h*0.72])
        X, Y, Z = interp_surface(arr)
        ax.imshow(Z.T, origin="lower", extent=[1,5,1,3], cmap=cmap, norm=norm,
                  interpolation="bicubic", aspect="auto")
        levels = np.linspace(vmin, vmax, 5)
        cs = ax.contour(X, Y, Z, levels=levels, colors="#123", linewidths=0.45, alpha=0.35)
        ax.clabel(cs, fontsize=6, inline=True, fmt="%.2f")
        # measured points
        for i in range(5):
            for j in range(3):
                ax.scatter(i+1, j+1, s=22, facecolor="white", edgecolor="black", linewidth=0.8, zorder=4)
        # min marker
        mn_idx = np.unravel_index(np.argmin(arr), arr.shape)
        ax.scatter(mn_idx[0]+1, mn_idx[1]+1, s=70, marker="s", facecolor="white",
                   edgecolor="black", linewidth=1.2, zorder=5)
        # centroid based on lower thickness distribution
        wear = arr.max() - arr
        total = wear.sum()
        Xp, Yp = np.meshgrid(np.arange(1,6), np.arange(1,4), indexing="ij")
        if total > 0:
            cx = (Xp*wear).sum()/total
            cy = (Yp*wear).sum()/total
        else:
            cx, cy = 3, 2
        ax.scatter(cx, cy, s=80, marker="o", facecolor="white", edgecolor="black", linewidth=2.0, zorder=6)
        ax.scatter(cx, cy, s=18, marker="o", facecolor="#0b2545", edgecolor="white", linewidth=0.5, zorder=7)
        ax.text(cx+0.12, cy+0.08, "Centroide", fontsize=6.5, fontweight="bold", color="#0b2545")
        ax.text(cx+0.12, cy-0.06, f"{Z.min():.2f}", fontsize=7.5, fontweight="bold", color="#0b55b8")

        ax.set_xlim(1,5)
        ax.set_ylim(1,3)
        ax.set_xticks([1,2,3,4,5])
        ax.set_yticks([1,2,3])
        ax.set_yticklabels(["A\nInterior","B\nCentro","C\nExterior"], fontsize=7)
        ax.set_xlabel("Posición circunferencial (1 → 5)", fontsize=7)
        ax.tick_params(axis="x", labelsize=7)
        for sp in ax.spines.values():
            sp.set_visible(False)

        # table
        ax_tbl = fig.add_axes([x+w*0.58, y+0.135, w*0.20, h*0.45])
        ax_tbl.axis("off")
        tbl_data = [["P","A","B","C"]] + [[str(i+1)] + [f"{arr[i,j]:.2f}" for j in range(3)] for i in range(5)]
        tbl = ax_tbl.table(cellText=tbl_data, cellLoc="center", loc="center")
        tbl.scale(1.0, 1.1)
        for (r,c), cell in tbl.get_celld().items():
            cell.set_edgecolor("#c5d0df")
            cell.set_linewidth(0.5)
            cell.get_text().set_fontsize(7)
            if r == 0:
                cell.set_facecolor("#0b2545")
                cell.get_text().set_color("white")
                cell.get_text().set_fontweight("bold")
        ax_tbl.set_title("LECTURA VISUAL  (mm)", fontsize=7.5, color="#0b2545", fontweight="bold", pad=3)

        # KPI
        k = get_metrics(arr)
        kpi_x = x+w*0.83
        kpi_y = y+0.10
        kpi_w = w*0.13
        kpi_h = h*0.65
        labels = [("PROMEDIO", k["mean"], "#0b2545"), ("MÍNIMO", k["min"], "red"),
                  ("MÁXIMO", k["max"], "#168323"), ("RANGO", k["range"], "#0b2545")]
        for idx, (lab, val, col) in enumerate(labels):
            yy = kpi_y + kpi_h - (idx+1)*0.055
            fig.add_artist(FancyBboxPatch((kpi_x, yy), kpi_w, 0.050, transform=fig.transFigure,
                                          boxstyle="round,pad=0.003,rounding_size=0.004",
                                          facecolor="#f8fafc", edgecolor="#d7e0ec", linewidth=0.6))
            fig.text(kpi_x+kpi_w/2, yy+0.033, lab, fontsize=6, color=col, ha="center", fontweight="bold")
            fig.text(kpi_x+kpi_w/2, yy+0.010, f"{val:.2f}", fontsize=10, color=col, ha="center", fontweight="bold")
            fig.text(kpi_x+kpi_w*0.78, yy+0.011, "mm", fontsize=5.5, color=col, ha="left", fontweight="bold")

        cond, cond_col = condition_from_values(arr)
        fig.add_artist(FancyBboxPatch((x+w*0.75, y+0.035), w*0.20, 0.045, transform=fig.transFigure,
                                      boxstyle="round,pad=0.004,rounding_size=0.006",
                                      facecolor=cond_col, edgecolor=cond_col, linewidth=0.7))
        fig.add_artist(Circle((x+w*0.775, y+0.057), 0.007, transform=fig.transFigure,
                              facecolor="none", edgecolor="white" if cond!="MONITOREAR" else "black", linewidth=1.0))
        fig.text(x+w*0.835, y+0.062, "CONDICIÓN", fontsize=6.2,
                 color="white" if cond!="MONITOREAR" else "#0b2545", ha="center", fontweight="bold")
        fig.text(x+w*0.835, y+0.045, cond, fontsize=9.0,
                 color="white" if cond!="MONITOREAR" else "#0b2545", ha="center", fontweight="bold")

    # Panels with highlighted borders by condition
    panel("A1 UPPER", data["A1 UPPER"], (0.025, 0.52, 0.47, 0.285), "#188038")
    panel("A1 LOWER", data["A1 LOWER"], (0.505, 0.52, 0.47, 0.285), "#b7791f")
    panel("B1 UPPER", data["B1 UPPER"], (0.025, 0.225, 0.47, 0.285), "#d7191c")
    panel("B1 LOWER", data["B1 LOWER"], (0.505, 0.225, 0.47, 0.285), "#b7791f")

    # Footer boxes
    footer_y = 0.065
    # scale
    fig.add_artist(FancyBboxPatch((0.025, footer_y), 0.29, 0.13, transform=fig.transFigure,
                                  boxstyle="round,pad=0.005,rounding_size=0.006",
                                  facecolor="white", edgecolor="#d7e0ec", linewidth=0.8))
    fig.text(0.055, footer_y+0.108, "ESCALA DE ESPESOR", fontsize=9, color="#0b2545", fontweight="bold")
    cax = fig.add_axes([0.045, footer_y+0.028, 0.018, 0.085])
    grad = np.linspace(vmin, vmax, 256).reshape(-1,1)
    cax.imshow(grad, cmap=cmap, norm=norm, aspect="auto", origin="lower")
    cax.set_xticks([])
    cax.set_yticks(np.linspace(0,255,5))
    cax.set_yticklabels([f"{v:.2f}" for v in np.linspace(vmin, vmax, 5)], fontsize=6)
    for sp in cax.spines.values():
        sp.set_visible(False)
    fig.text(0.085, footer_y+0.100, "MAYOR ESPESOR", fontsize=7, color="#168323", fontweight="bold")
    fig.text(0.085, footer_y+0.030, "MENOR ESPESOR", fontsize=7, color="red", fontweight="bold")

    # symbol
    fig.text(0.195, footer_y+0.108, "SIMBOLOGÍA", fontsize=9, color="#0b2545", fontweight="bold")
    sym_y = [footer_y+0.087, footer_y+0.067, footer_y+0.047, footer_y+0.027]
    sym_txt = ["Punto de medición", "Menor espesor", "Centroide de desgaste", "Curva de nivel"]
    for i, txt in enumerate(sym_txt):
        fig.text(0.195, sym_y[i], "○" if i==0 else ("□" if i==1 else ("◎" if i==2 else "—")),
                 fontsize=10, color="black")
        fig.text(0.215, sym_y[i]+0.002, txt, fontsize=7.5, color="#0b2545")

    # risk summary
    fig.add_artist(FancyBboxPatch((0.325, footer_y), 0.245, 0.13, transform=fig.transFigure,
                                  boxstyle="round,pad=0.005,rounding_size=0.006",
                                  facecolor="white", edgecolor="#d7e0ec", linewidth=0.8))
    fig.text(0.395, footer_y+0.108, "MAPA DE RIESGO (RESUMEN)", fontsize=9, color="#0b2545", fontweight="bold")
    risk_items = [("A1\nUPPER", "#228b22"), ("A1\nLOWER", "#f5b400"), ("B1\nUPPER", "#e10600"), ("B1\nLOWER", "#f5b400")]
    for i, (lab, col) in enumerate(risk_items):
        x = 0.345 + i*0.052
        fig.text(x, footer_y+0.072, lab, fontsize=7, color="#0b2545", ha="center", fontweight="bold")
        fig.add_artist(Circle((x, footer_y+0.033), 0.008, transform=fig.transFigure,
                              facecolor=col, edgecolor="black", linewidth=0.7))

    # diagnosis
    fig.add_artist(FancyBboxPatch((0.58, footer_y), 0.395, 0.13, transform=fig.transFigure,
                                  boxstyle="round,pad=0.005,rounding_size=0.006",
                                  facecolor="white", edgecolor="#d7e0ec", linewidth=0.8))
    fig.text(0.72, footer_y+0.108, "DIAGNÓSTICO RESUMEN", fontsize=9, color="#0b2545", fontweight="bold")
    lines = [
        ("●", "#228b22", "A1 Upper: Desgaste uniforme. Sin anomalías relevantes."),
        ("●", "#e10600", "B1 Upper: Desgaste localizado severo. Recomendada inspección detallada."),
        ("●", "#f5b400", "B1 Lower: Desgaste moderado con tendencia uniforme. Monitorear."),
    ]
    for i, (icon, col, txt) in enumerate(lines):
        yy = footer_y+0.083 - i*0.033
        fig.text(0.61, yy, icon, fontsize=17, color=col, va="center")
        fig.text(0.635, yy, txt, fontsize=8.2, color="#0b2545", va="center")

    fig.text(0.025, 0.025, "NOTA: Medidas en mm. Valores interpolados como apoyo visual; los puntos de medición representan los datos reales.",
             fontsize=7.5, color="#0b2545")
    fig.text(0.895, 0.025, "LUFUSSA   |   CONFIDENCIAL", fontsize=8, color="#65758b")

    return fig

# =========================
# UI
# =========================
st.markdown("""
<div class="main-title">
    <div class="title-text">OEM Insight Studio Web v1.0</div>
    <div class="subtitle-text">Generador web de reporte visual para Big End Bearing - WSM</div>
</div>
""", unsafe_allow_html=True)

left, right = st.columns([1, 1])

with left:
    st.subheader("1. Cargar archivos Excel")
    banco_a = st.file_uploader("Archivo Banco A", type=["xlsx"], key="a")
    banco_b = st.file_uploader("Archivo Banco B", type=["xlsx"], key="b")

with right:
    st.subheader("2. Datos del reporte")
    unidad = st.text_input("Equipo / Unidad", "GENSET 11")
    motor = st.text_input("Motor", "WÄRTSILÄ 46")
    horas = st.text_input("Horas de operación", "12,845 h")
    fecha = st.text_input("Fecha", date.today().strftime("%d / %m / %Y"))
    reporte = st.text_input("Reporte No.", "WSM-2026-0702")
    inspector = st.text_input("Inspector", "")

st.subheader("3. Selección de cilindros")
c1, c2 = st.columns(2)
with c1:
    cil_a = st.text_input("Cilindro Banco A", "A1")
with c2:
    cil_b = st.text_input("Cilindro Banco B", "B1")

if st.button("Generar reporte visual", type="primary"):
    if banco_a is None or banco_b is None:
        st.warning("Carga ambos archivos Excel: Banco A y Banco B.")
    else:
        try:
            sheet_a = find_sheet(banco_a, cil_a)
            sheet_b = find_sheet(banco_b, cil_b)

            # Reset file pointers after reading sheetnames
            banco_a.seek(0)
            banco_b.seek(0)
            arr_a_u = read_bearing(banco_a, sheet_a, True)
            banco_a.seek(0)
            arr_a_l = read_bearing(banco_a, sheet_a, False)
            banco_b.seek(0)
            arr_b_u = read_bearing(banco_b, sheet_b, True)
            banco_b.seek(0)
            arr_b_l = read_bearing(banco_b, sheet_b, False)

            report_data = {
                f"{cil_a} UPPER": arr_a_u,
                f"{cil_a} LOWER": arr_a_l,
                f"{cil_b} UPPER": arr_b_u,
                f"{cil_b} LOWER": arr_b_l,
            }

            # Convert keys expected by build_report for v1
            std_data = {
                "A1 UPPER": arr_a_u,
                "A1 LOWER": arr_a_l,
                "B1 UPPER": arr_b_u,
                "B1 LOWER": arr_b_l,
            }

            meta = {
                "unidad": unidad,
                "motor": motor,
                "horas": horas,
                "fecha": fecha,
                "reporte": reporte,
                "inspector": inspector,
                "cil_a": cil_a,
                "cil_b": cil_b,
            }

            fig = build_report(std_data, meta)

            png_buffer = BytesIO()
            pdf_buffer = BytesIO()
            fig.savefig(png_buffer, format="png", bbox_inches="tight", facecolor="white")
            fig.savefig(pdf_buffer, format="pdf", bbox_inches="tight", facecolor="white")
            png_buffer.seek(0)
            pdf_buffer.seek(0)
            plt.close(fig)

            st.success("Reporte generado correctamente.")
            st.image(png_buffer, use_container_width=True)

            st.download_button(
                "Descargar PNG",
                data=png_buffer.getvalue(),
                file_name="Reporte_Big_End_Bearing_WSM.png",
                mime="image/png"
            )
            st.download_button(
                "Descargar PDF",
                data=pdf_buffer.getvalue(),
                file_name="Reporte_Big_End_Bearing_WSM.pdf",
                mime="application/pdf"
            )

        except Exception as e:
            st.error(f"No se pudo generar el reporte: {e}")

st.markdown('<p class="small-note">Versión piloto: estructura optimizada para archivos Excel con hojas tipo "BEB A1", "BEB B1" y tablas Upper/Lower en posiciones estándar.</p>', unsafe_allow_html=True)
