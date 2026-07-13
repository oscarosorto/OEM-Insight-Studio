
import streamlit as st
import numpy as np
import openpyxl
import matplotlib.pyplot as plt
from matplotlib.colors import ListedColormap, BoundaryNorm
from io import BytesIO

st.set_page_config(
    page_title="Gráfico Simple de Cojinete",
    page_icon="📊",
    layout="wide"
)

st.title("📊 Medición de Cojinetes Big End Bearing")
st.caption("Carga el archivo Excel una sola vez y selecciona banco, cilindro y tipo de cojinete.")

# ============================
# DATOS GENERALES
# ============================
st.sidebar.header("Datos generales")
equipo = st.sidebar.text_input("Equipo / Unidad", "GENSET 11")
horas = st.sidebar.text_input("Horas de operación", "12000 h")
supervisor = st.sidebar.text_input("Supervisor", "")

archivo = st.file_uploader("Cargar archivo Excel", type=["xlsx"])

# ============================
# LECTURA EXCLUSIVA DE mediciones BEB
# ============================
def leer_matriz_mediciones_beb(file, banco, cilindro, tipo):
    """
    Lee únicamente la hoja 'mediciones BEB'.

    Distribución:
    - Cada cilindro ocupa un bloque de 10 filas.
    - Cilindro 1: datos en filas 6-10
    - Cilindro 2: datos en filas 16-20
    - ...
    - Cilindro 9: datos en filas 86-90

    Columnas:
    Banco A UPPER: C:E
    Banco A LOWER: G:I
    Banco B UPPER: L:N
    Banco B LOWER: P:R
    """
    file.seek(0)
    wb = openpyxl.load_workbook(file, data_only=True)

    nombre_hoja = "mediciones BEB"
    if nombre_hoja not in wb.sheetnames:
        raise ValueError("El archivo no contiene la hoja 'mediciones BEB'.")

    ws = wb[nombre_hoja]

    fila_inicial = 6 + (int(cilindro) - 1) * 10
    filas = range(fila_inicial, fila_inicial + 5)

    columnas_por_seleccion = {
        ("A", "UPPER"): [3, 4, 5],     # C:E
        ("A", "LOWER"): [7, 8, 9],     # G:I
        ("B", "UPPER"): [12, 13, 14],  # L:N
        ("B", "LOWER"): [16, 17, 18],  # P:R
    }

    columnas = columnas_por_seleccion[(banco, tipo)]

    data = []
    for r in filas:
        row = []
        for c in columnas:
            value = ws.cell(r, c).value
            if value is None:
                celda = ws.cell(r, c).coordinate
                raise ValueError(
                    f"No existe una medición en {celda} para "
                    f"{banco}{cilindro} {tipo}."
                )
            try:
                row.append(float(value))
            except (TypeError, ValueError):
                celda = ws.cell(r, c).coordinate
                raise ValueError(f"El valor de {celda} no es numérico.")
        data.append(row)

    return np.array(data, dtype=float)

def interp_surface(arr, nx=220, ny=150):
    x = np.arange(1, 6)
    y = np.arange(1, 4)
    xi = np.linspace(1, 5, nx)
    temp = np.array([np.interp(xi, x, arr[:, j]) for j in range(3)])
    yi = np.linspace(1, 3, ny)
    zi = np.array([np.interp(yi, y, temp[:, i]) for i in range(nx)]).T
    X, Y = np.meshgrid(xi, yi)
    return X, Y, zi

def wear_centroid(arr):
    wear = arr.max() - arr
    total = wear.sum()
    Xp, Yp = np.meshgrid(np.arange(1, 6), np.arange(1, 4), indexing="ij")
    if total <= 0:
        return 3, 2
    return (Xp * wear).sum() / total, (Yp * wear).sum() / total

def condition(arr):
    minimo = float(arr.min())

    if minimo < 0.90:
        return "REVISAR", "#d7191c"
    elif minimo < 1.00:
        return "DESGASTE IMPORTANTE", "#f57c00"
    elif minimo < 1.10:
        return "SEGUIMIENTO", "#ffd54f"
    elif minimo < 1.20:
        return "NORMAL", "#7bc96f"
    return "EXCELENTE", "#228b22"

def crear_grafico(arr, equipo, banco, cilindro, tipo, horas, supervisor):
    # Escala técnica fija ya aprobada
    bounds = [0.00, 0.90, 1.00, 1.10, 1.20, 1.40]

    cmap = ListedColormap([
        "#b30000",  # <0.90 desgaste severo
        "#f57c00",  # 0.90-0.99 desgaste importante
        "#ffd54f",  # 1.00-1.09 seguimiento
        "#7bc96f",  # 1.10-1.19 normal
        "#1b8f3a"   # >=1.20 excelente
    ])

    norm = BoundaryNorm(bounds, cmap.N)
    vmin = bounds[0]
    vmax = bounds[-1]

    fig = plt.figure(figsize=(13, 8), dpi=180)
    fig.patch.set_facecolor("white")

    identificacion = f"{banco}{cilindro}"

    fig.text(
        0.04, 0.94,
        "MEDICIÓN DE COJINETES BIG END BEARING",
        fontsize=18,
        fontweight="bold",
        color="#0b2545"
    )
    fig.text(
        0.04, 0.905,
        f"Equipo: {equipo}   |   Cilindro: {identificacion}   |   Tipo: {tipo}",
        fontsize=10,
        color="#0b55b8"
    )
    fig.text(
        0.04, 0.88,
        f"Horas de operación: {horas}   |   Supervisor: {supervisor}",
        fontsize=10,
        color="#0b2545"
    )

    # Mapa visual
    ax = fig.add_axes([0.07, 0.22, 0.55, 0.58])
    X, Y, Z = interp_surface(arr)

    ax.imshow(
        Z.T,
        origin="lower",
        extent=[1, 5, 1, 3],
        cmap=cmap,
        norm=norm,
        interpolation="bicubic",
        aspect="auto"
    )

    # Curvas basadas en límites técnicos útiles
    niveles_contorno = [0.90, 1.00, 1.10, 1.20]
    zmin, zmax = float(Z.min()), float(Z.max())
    niveles_visibles = [n for n in niveles_contorno if zmin < n < zmax]

    if niveles_visibles:
        cs = ax.contour(
            X, Y, Z,
            levels=niveles_visibles,
            colors="#0b2545",
            linewidths=0.55,
            alpha=0.40
        )
        ax.clabel(cs, fontsize=7, inline=True, fmt="%.2f")

    for i in range(5):
        for j in range(3):
            ax.scatter(
                i + 1, j + 1,
                s=45,
                facecolor="white",
                edgecolor="black",
                linewidth=0.9,
                zorder=4
            )
            ax.text(
                i + 1, j + 1,
                f"{arr[i, j]:.2f}",
                ha="center",
                va="center",
                fontsize=7,
                color="#0b2545",
                zorder=5
            )

    min_idx = np.unravel_index(np.argmin(arr), arr.shape)
    ax.scatter(
        min_idx[0] + 1,
        min_idx[1] + 1,
        s=120,
        marker="s",
        facecolor="none",
        edgecolor="black",
        linewidth=1.6,
        zorder=6
    )

    cx, cy = wear_centroid(arr)
    ax.scatter(
        cx, cy,
        s=110,
        facecolor="white",
        edgecolor="black",
        linewidth=2.0,
        zorder=7
    )
    ax.scatter(
        cx, cy,
        s=25,
        facecolor="#0b2545",
        edgecolor="white",
        linewidth=0.6,
        zorder=8
    )
    ax.text(
        cx + 0.12,
        cy + 0.08,
        "Centroide",
        fontsize=8,
        fontweight="bold",
        color="#0b2545"
    )

    ax.set_xlim(1, 5)
    ax.set_ylim(1, 3)
    ax.set_xticks([1, 2, 3, 4, 5])
    ax.set_yticks([1, 2, 3])
    ax.set_yticklabels(
        ["A\nInterior", "B\nCentro", "C\nExterior"],
        fontsize=9
    )
    ax.set_xlabel("Posición circunferencial 1 → 5", fontsize=10)

    for sp in ax.spines.values():
        sp.set_visible(False)

    # Tabla
    ax_tbl = fig.add_axes([0.68, 0.48, 0.25, 0.28])
    ax_tbl.axis("off")

    table_data = [["P", "A", "B", "C"]]
    for i in range(5):
        table_data.append(
            [str(i + 1)] + [f"{arr[i, j]:.2f}" for j in range(3)]
        )

    tbl = ax_tbl.table(
        cellText=table_data,
        loc="center",
        cellLoc="center"
    )
    tbl.auto_set_font_size(False)
    tbl.set_fontsize(9)
    tbl.scale(1.1, 1.3)

    for (r, c), cell in tbl.get_celld().items():
        cell.set_edgecolor("#c5d0df")
        cell.set_linewidth(0.5)
        if r == 0:
            cell.set_facecolor("#0b2545")
            cell.get_text().set_color("white")
            cell.get_text().set_fontweight("bold")

    ax_tbl.set_title(
        "LECTURA VISUAL (mm)",
        fontsize=10,
        color="#0b2545",
        fontweight="bold"
    )

    # KPIs
    promedio = float(arr.mean())
    minimo = float(arr.min())
    maximo = float(arr.max())
    rango = maximo - minimo
    estado, color_estado = condition(arr)

    ax_kpi = fig.add_axes([0.68, 0.23, 0.25, 0.18])
    ax_kpi.axis("off")

    kpis = [
        ("PROMEDIO", promedio, "#0b2545"),
        ("MÍNIMO", minimo, "red"),
        ("MÁXIMO", maximo, "#168323"),
        ("RANGO", rango, "#0b2545")
    ]

    for idx, (label, value, color) in enumerate(kpis):
        x = (idx % 2) * 0.50
        y = 0.63 if idx < 2 else 0.18
        ax_kpi.text(
            x + 0.23, y + 0.16,
            label,
            ha="center",
            fontsize=8,
            color=color,
            fontweight="bold"
        )
        ax_kpi.text(
            x + 0.23, y,
            f"{value:.2f} mm",
            ha="center",
            fontsize=13,
            color=color,
            fontweight="bold"
        )

    fig.text(
        0.72, 0.14,
        f"CONDICIÓN: {estado}",
        fontsize=16,
        color=color_estado,
        fontweight="bold"
    )

    # Escala técnica fija
    ax_scale = fig.add_axes([0.08, 0.08, 0.42, 0.045])
    grad = np.linspace(vmin, vmax, 256).reshape(1, -1)

    ax_scale.imshow(
        grad,
        cmap=cmap,
        norm=norm,
        aspect="auto",
        extent=[vmin, vmax, 0, 1]
    )

    ax_scale.set_yticks([])
    ax_scale.set_xticks([0.90, 1.00, 1.10, 1.20, 1.40])
    ax_scale.set_xticklabels(
        ["0.90", "1.00", "1.10", "1.20", "≥1.40"],
        fontsize=8
    )

    for sp in ax_scale.spines.values():
        sp.set_visible(False)

    fig.text(
        0.08, 0.135,
        "<0.90 SEVERO",
        fontsize=8,
        color="#b30000",
        fontweight="bold"
    )
    fig.text(
        0.19, 0.135,
        "0.90–0.99 DESGASTE",
        fontsize=8,
        color="#f57c00",
        fontweight="bold"
    )
    fig.text(
        0.34, 0.135,
        "≥1.20 NORMAL / EXCELENTE",
        fontsize=8,
        color="#168323",
        fontweight="bold"
    )

    return fig


def crear_grafico_doble(arr_upper, arr_lower, equipo, banco, cilindro, horas, supervisor):
    bounds = [0.00, 0.90, 1.00, 1.10, 1.20, 1.40]
    cmap = ListedColormap([
        "#b30000",
        "#f57c00",
        "#ffd54f",
        "#7bc96f",
        "#1b8f3a"
    ])
    norm = BoundaryNorm(bounds, cmap.N)

    fig = plt.figure(figsize=(14, 10), dpi=180)
    fig.patch.set_facecolor("white")

    identificacion = f"{banco}{cilindro}"

    fig.text(
        0.04, 0.955,
        "MEDICIÓN DE COJINETES BIG END BEARING",
        fontsize=19,
        fontweight="bold",
        color="#0b2545"
    )
    fig.text(
        0.04, 0.925,
        f"Equipo: {equipo}   |   Cilindro: {identificacion}   |   UPPER y LOWER",
        fontsize=10.5,
        color="#0b55b8"
    )
    fig.text(
        0.04, 0.900,
        f"Horas de operación: {horas}   |   Supervisor: {supervisor}",
        fontsize=10,
        color="#0b2545"
    )

    def draw_section(arr, title, y0):
        X, Y, Z = interp_surface(arr)

        ax = fig.add_axes([0.06, y0, 0.50, 0.30])
        ax.imshow(
            Z.T,
            origin="lower",
            extent=[1, 5, 1, 3],
            cmap=cmap,
            norm=norm,
            interpolation="bicubic",
            aspect="auto"
        )

        niveles = [n for n in [0.90, 1.00, 1.10, 1.20] if float(Z.min()) < n < float(Z.max())]
        if niveles:
            cs = ax.contour(
                X, Y, Z,
                levels=niveles,
                colors="#0b2545",
                linewidths=0.55,
                alpha=0.40
            )
            ax.clabel(cs, fontsize=7, inline=True, fmt="%.2f")

        for i in range(5):
            for j in range(3):
                ax.scatter(
                    i + 1, j + 1,
                    s=42,
                    facecolor="white",
                    edgecolor="black",
                    linewidth=0.85,
                    zorder=4
                )
                ax.text(
                    i + 1, j + 1,
                    f"{arr[i, j]:.2f}",
                    ha="center",
                    va="center",
                    fontsize=7,
                    color="#0b2545",
                    zorder=5
                )

        min_idx = np.unravel_index(np.argmin(arr), arr.shape)
        ax.scatter(
            min_idx[0] + 1,
            min_idx[1] + 1,
            s=110,
            marker="s",
            facecolor="none",
            edgecolor="black",
            linewidth=1.5,
            zorder=6
        )

        cx, cy = wear_centroid(arr)
        ax.scatter(
            cx, cy,
            s=95,
            facecolor="white",
            edgecolor="black",
            linewidth=1.8,
            zorder=7
        )
        ax.scatter(
            cx, cy,
            s=22,
            facecolor="#0b2545",
            edgecolor="white",
            linewidth=0.5,
            zorder=8
        )
        ax.text(
            cx + 0.12,
            cy + 0.08,
            "Centroide",
            fontsize=7.5,
            fontweight="bold",
            color="#0b2545"
        )

        ax.set_title(title, loc="left", fontsize=14, fontweight="bold", color="#0b2545")
        ax.set_xlim(1, 5)
        ax.set_ylim(1, 3)
        ax.set_xticks([1, 2, 3, 4, 5])
        ax.set_yticks([1, 2, 3])
        ax.set_yticklabels(["A\nInterior", "B\nCentro", "C\nExterior"], fontsize=8)
        ax.set_xlabel("Posición circunferencial 1 → 5", fontsize=9)

        for sp in ax.spines.values():
            sp.set_visible(False)

        ax_tbl = fig.add_axes([0.61, y0 + 0.06, 0.18, 0.20])
        ax_tbl.axis("off")

        table_data = [["P", "A", "B", "C"]]
        for i in range(5):
            table_data.append(
                [str(i + 1)] + [f"{arr[i, j]:.2f}" for j in range(3)]
            )

        tbl = ax_tbl.table(
            cellText=table_data,
            loc="center",
            cellLoc="center"
        )
        tbl.auto_set_font_size(False)
        tbl.set_fontsize(8)
        tbl.scale(1.05, 1.22)

        for (r, c), cell in tbl.get_celld().items():
            cell.set_edgecolor("#c5d0df")
            cell.set_linewidth(0.5)
            if r == 0:
                cell.set_facecolor("#0b2545")
                cell.get_text().set_color("white")
                cell.get_text().set_fontweight("bold")

        ax_tbl.set_title("LECTURA VISUAL (mm)", fontsize=9, color="#0b2545", fontweight="bold")

        promedio = float(arr.mean())
        minimo = float(arr.min())
        maximo = float(arr.max())
        rango = maximo - minimo
        estado, color_estado = condition(arr)

        ax_kpi = fig.add_axes([0.82, y0 + 0.04, 0.14, 0.24])
        ax_kpi.axis("off")

        kpis = [
            ("PROMEDIO", promedio, "#0b2545"),
            ("MÍNIMO", minimo, "red"),
            ("MÁXIMO", maximo, "#168323"),
            ("RANGO", rango, "#0b2545")
        ]

        yy = 0.92
        for label, value, color in kpis:
            ax_kpi.text(
                0.5, yy,
                label,
                ha="center",
                fontsize=7,
                color=color,
                fontweight="bold"
            )
            ax_kpi.text(
                0.5, yy - 0.11,
                f"{value:.2f} mm",
                ha="center",
                fontsize=11,
                color=color,
                fontweight="bold"
            )
            yy -= 0.22

        ax_kpi.text(
            0.5, 0.02,
            estado,
            ha="center",
            fontsize=11,
            color=color_estado,
            fontweight="bold"
        )

    draw_section(arr_upper, "UPPER", 0.53)
    draw_section(arr_lower, "LOWER", 0.17)

    ax_scale = fig.add_axes([0.12, 0.07, 0.76, 0.035])
    grad = np.linspace(0.00, 1.40, 256).reshape(1, -1)
    ax_scale.imshow(
        grad,
        cmap=cmap,
        norm=norm,
        aspect="auto",
        extent=[0.00, 1.40, 0, 1]
    )
    ax_scale.set_yticks([])
    ax_scale.set_xticks([0.90, 1.00, 1.10, 1.20, 1.40])
    ax_scale.set_xticklabels(["0.90", "1.00", "1.10", "1.20", "≥1.40"], fontsize=8)
    for sp in ax_scale.spines.values():
        sp.set_visible(False)

    fig.text(0.12, 0.115, "<0.90 SEVERO", fontsize=8, color="#b30000", fontweight="bold")
    fig.text(0.29, 0.115, "0.90–0.99 DESGASTE", fontsize=8, color="#f57c00", fontweight="bold")
    fig.text(0.55, 0.115, "1.00–1.09 SEGUIMIENTO", fontsize=8, color="#c49a00", fontweight="bold")
    fig.text(0.77, 0.115, "≥1.10 NORMAL / EXCELENTE", fontsize=8, color="#168323", fontweight="bold")

    return fig


# ============================
# INTERFAZ DE SELECCIÓN
# ============================
if archivo:
    st.success("Archivo cargado. Fuente de datos: hoja 'mediciones BEB'.")

    c1, c2 = st.columns(2)

    with c1:
        banco = st.selectbox("Banco", ["A", "B"])

    with c2:
        cilindro = st.selectbox(
            "Cilindro",
            list(range(1, 10))
        )

    st.write(
        f"**Selección actual:** Cilindro {banco}{cilindro} — UPPER y LOWER"
    )

    if st.button("Generar gráfico", type="primary"):
        try:
            arr_upper = leer_matriz_mediciones_beb(
                archivo, banco, cilindro, "UPPER"
            )
            arr_lower = leer_matriz_mediciones_beb(
                archivo, banco, cilindro, "LOWER"
            )

            fig = crear_grafico_doble(
                arr_upper,
                arr_lower,
                equipo,
                banco,
                cilindro,
                horas,
                supervisor
            )

            png = BytesIO()
            pdf = BytesIO()

            fig.savefig(
                png,
                format="png",
                bbox_inches="tight",
                facecolor="white"
            )
            fig.savefig(
                pdf,
                format="pdf",
                bbox_inches="tight",
                facecolor="white"
            )

            png.seek(0)
            pdf.seek(0)

            plt.close(fig)

            st.success("Gráfico generado correctamente.")
            st.image(png, use_container_width=True)

            nombre_base = f"grafico_{banco}{cilindro}_UPPER_LOWER"

            st.download_button(
                "Descargar PNG",
                data=png.getvalue(),
                file_name=f"{nombre_base}.png",
                mime="image/png"
            )

            st.download_button(
                "Descargar PDF",
                data=pdf.getvalue(),
                file_name=f"{nombre_base}.pdf",
                mime="application/pdf"
            )

        except Exception as e:
            st.error(f"No se pudo generar el gráfico: {e}")
else:
    st.info("Carga el archivo Excel para comenzar.")
