import streamlit as st
import numpy as np
import openpyxl
import matplotlib.pyplot as plt
from matplotlib.colors import LinearSegmentedColormap, Normalize
from matplotlib.patches import FancyBboxPatch, Circle
from io import BytesIO
from datetime import date
from pathlib import Path

st.set_page_config(page_title='OEM Insight Studio', page_icon='📊', layout='wide')
st.markdown('''<style>.stApp{background:#f3f6fb}.box{background:white;padding:18px;border-radius:18px;border:1px solid #d9e2ef}.title{color:#0b2545;font-size:32px;font-weight:800}.sub{color:#31517a;font-size:16px}</style>''', unsafe_allow_html=True)

def load_wb(f):
    f.seek(0); return openpyxl.load_workbook(f, data_only=True)
def sheets(f): return load_wb(f).sheetnames

def read_template(f, sheet, upper=True):
    ws=load_wb(f)[sheet]
    rows=range(14,19); cols=[7,8,9] if upper else [18,19,20]
    out=[]
    for r in rows:
        row=[]
        for c in cols:
            v=ws.cell(r,c).value
            if v is None: raise ValueError(f'Celda vacía: {sheet} fila {r} columna {c}')
            row.append(float(v))
        out.append(row)
    return np.array(out,float)

def read_manual(f,sheet,r0,c0):
    ws=load_wb(f)[sheet]; out=[]
    for r in range(r0,r0+5):
        out.append([float(ws.cell(r,c).value) for c in range(c0,c0+3)])
    return np.array(out,float)

def cond(arr):
    mn=float(arr.min()); rg=float(arr.max()-arr.min())
    if mn<0.85 or rg>0.50: return 'REVISAR','#e10600','#d7191c'
    if mn<1.05 or rg>0.28: return 'MONITOREAR','#f5b400','#b7791f'
    return 'NORMAL','#228b22','#188038'

def interp(arr,nx=220,ny=150):
    x=np.arange(1,6); y=np.arange(1,4); xi=np.linspace(1,5,nx)
    temp=np.array([np.interp(xi,x,arr[:,j]) for j in range(3)])
    yi=np.linspace(1,3,ny); zi=np.array([np.interp(yi,y,temp[:,i]) for i in range(nx)]).T
    X,Y=np.meshgrid(xi,yi); return X,Y,zi

def centroid(arr):
    wear=arr.max()-arr; total=wear.sum(); Xp,Yp=np.meshgrid(np.arange(1,6),np.arange(1,4),indexing='ij')
    if total<=0: return 3,2
    return (Xp*wear).sum()/total,(Yp*wear).sum()/total

def create_report(data,meta):
    cmap=LinearSegmentedColormap.from_list('wear',['#b30000','#e34a33','#fdbb84','#fee08b','#d9ef8b','#66bd63','#1a9850'])
    vals=np.concatenate([v.flatten() for v in data.values()])
    vmin=np.floor((vals.min()-0.02)*100)/100; vmax=np.ceil((vals.max()+0.02)*100)/100; norm=Normalize(vmin=vmin,vmax=vmax)
    fig=plt.figure(figsize=(16,10),dpi=180); fig.patch.set_facecolor('white')
    ax_logo=fig.add_axes([0.025,0.905,0.145,0.08]); ax_logo.axis('off')
    logo=Path('assets/lufussa_logo.png')
    if logo.exists(): ax_logo.imshow(plt.imread(str(logo)))
    else: ax_logo.text(0,0.5,'LUFUSSA',fontsize=22,fontweight='bold',color='#5b6570')
    fig.text(0.205,0.955,'INSPECCIÓN DE COJINETES BIG END BEARING – FIRMA SUPERFICIAL DE DESGASTE',fontsize=17,fontweight='bold',color='#0b2545',ha='left',va='center')
    fig.text(0.285,0.925,f"Cilindros {meta['cil_a']} y {meta['cil_b']}   |   Datos reales de medición   |   Rojo = menor espesor   /   Verde = mayor espesor",fontsize=11,color='#0b55b8',ha='left',va='center')
    axm=fig.add_axes([0.875,0.885,0.105,0.105]); axm.axis('off'); axm.add_patch(FancyBboxPatch((0,0),1,1,boxstyle='round,pad=0.015',facecolor='white',edgecolor='#c9d4e4'))
    axm.text(.08,.78,'FECHA:',fontsize=7,fontweight='bold',color='#0b2545'); axm.text(.58,.78,meta['fecha'],fontsize=7,fontweight='bold',ha='center',color='#0b2545')
    axm.text(.08,.50,'REPORTE No.:',fontsize=7,fontweight='bold',color='#0b2545'); axm.text(.60,.50,meta['reporte'],fontsize=7,fontweight='bold',ha='center',color='#0b2545')
    axm.text(.08,.21,'INSPECTOR:',fontsize=7,fontweight='bold',color='#0b2545'); axm.text(.60,.21,meta['inspector'] or '__________',fontsize=7,ha='center',color='#0b2545')
    axe=fig.add_axes([0.025,0.815,0.830,0.060]); axe.axis('off'); axe.add_patch(FancyBboxPatch((0,0),1,1,boxstyle='round,pad=0.008',facecolor='white',edgecolor='#c9d4e4'))
    fields=[('EQUIPO / UNIDAD',meta['unidad'],.17,1),('MOTOR',meta['motor'],.21,0),('BANCO','A y B',.15,0),('HORAS DE OPERACIÓN',meta['horas'],.22,0),('TIPO DE COJINETE','BIG END BEARING',.25,0)]
    x=.015
    for lab,val,w,dark in fields:
        if dark:
            axe.add_patch(FancyBboxPatch((x,.12),w,.76,boxstyle='round,pad=0.006',facecolor='#0b2545',edgecolor='#0b2545')); col='white'
        else:
            axe.add_line(plt.Line2D([x,x],[.15,.85],color='#c9d4e4')); col='#0b2545'
        axe.text(x+w/2,.62,lab,ha='center',va='center',fontsize=7,color=col,fontweight='bold'); axe.text(x+w/2,.34,val,ha='center',va='center',fontsize=10.5,color=col,fontweight='bold'); x+=w
    def panel(title,arr,x,y,w,h):
        cl,cc,bc=cond(arr)
        axb=fig.add_axes([x,y,w,h]); axb.axis('off'); axb.add_patch(FancyBboxPatch((0,0),1,1,boxstyle='round,pad=0.006',facecolor='white',edgecolor=bc,linewidth=1.0)); axb.text(.045,.94,title,fontsize=13,color=bc,fontweight='bold',va='top')
        ax=fig.add_axes([x+w*.055,y+h*.165,w*.455,h*.675]); X,Y,Z=interp(arr)
        ax.imshow(Z.T,origin='lower',extent=[1,5,1,3],cmap=cmap,norm=norm,interpolation='bicubic',aspect='auto')
        cs=ax.contour(X,Y,Z,levels=np.linspace(vmin,vmax,5),colors='#0b2545',linewidths=.45,alpha=.34); ax.clabel(cs,fontsize=6,inline=True,fmt='%.2f')
        for i in range(5):
            for j in range(3): ax.scatter(i+1,j+1,s=22,facecolor='white',edgecolor='black',linewidth=.75,zorder=4)
        mi=np.unravel_index(np.argmin(arr),arr.shape); ax.scatter(mi[0]+1,mi[1]+1,s=70,marker='s',facecolor='white',edgecolor='black',linewidth=1.2,zorder=5)
        cx,cy=centroid(arr); ax.scatter(cx,cy,s=80,facecolor='white',edgecolor='black',linewidth=2,zorder=6); ax.scatter(cx,cy,s=18,facecolor='#0b2545',edgecolor='white',linewidth=.5,zorder=7); ax.text(cx+.12,cy+.08,'Centroide',fontsize=6.5,fontweight='bold',color='#0b2545')
        ax.set_xlim(1,5); ax.set_ylim(1,3); ax.set_xticks([1,2,3,4,5]); ax.set_yticks([1,2,3]); ax.set_yticklabels(['A\nInterior','B\nCentro','C\nExterior'],fontsize=7); ax.set_xlabel('Posición circunferencial (1 → 5)',fontsize=7); ax.tick_params(axis='x',labelsize=7); [sp.set_visible(False) for sp in ax.spines.values()]
        axt=fig.add_axes([x+w*.565,y+h*.280,w*.205,h*.405]); axt.axis('off'); tbl=[['P','A','B','C']]+[[str(i+1)]+[f'{arr[i,j]:.2f}' for j in range(3)] for i in range(5)]
        table=axt.table(cellText=tbl,loc='center',cellLoc='center'); table.auto_set_font_size(False); table.set_fontsize(6.8); table.scale(1,1.2)
        for (r,c),cell in table.get_celld().items():
            cell.set_edgecolor('#c5d0df'); cell.set_linewidth(.45)
            if r==0: cell.set_facecolor('#0b2545'); cell.get_text().set_color('white'); cell.get_text().set_fontweight('bold')
        axt.set_title('LECTURA VISUAL  (mm)',fontsize=7.5,color='#0b2545',fontweight='bold',pad=3)
        k={'PROMEDIO':arr.mean(),'MÍNIMO':arr.min(),'MÁXIMO':arr.max(),'RANGO':arr.max()-arr.min()}; cols={'PROMEDIO':'#0b2545','MÍNIMO':'red','MÁXIMO':'#168323','RANGO':'#0b2545'}
        kx=x+w*.805; ky=y+h*.250; kw=w*.145; kh=h*.565
        for idx,(lab,val) in enumerate(k.items()):
            yy=ky+kh-(idx+1)*h*.145; fig.add_artist(FancyBboxPatch((kx,yy),kw,h*.120,transform=fig.transFigure,boxstyle='round,pad=0.003,rounding_size=0.004',facecolor='#f8fafc',edgecolor='#d7e0ec',linewidth=.55))
            fig.text(kx+kw/2,yy+h*.078,lab,fontsize=5.8,color=cols[lab],ha='center',fontweight='bold'); fig.text(kx+kw/2,yy+h*.028,f'{val:.2f}',fontsize=10,color=cols[lab],ha='center',fontweight='bold'); fig.text(kx+kw*.78,yy+h*.032,'mm',fontsize=5.2,color=cols[lab],ha='left',fontweight='bold')
        cx0=x+w*.735; cy0=y+h*.045; cw=w*.225; ch=h*.125; fig.add_artist(FancyBboxPatch((cx0,cy0),cw,ch,transform=fig.transFigure,boxstyle='round,pad=0.004,rounding_size=0.006',facecolor=cc,edgecolor=cc,linewidth=.7)); txt='white' if cl!='MONITOREAR' else '#0b2545'
        fig.add_artist(Circle((cx0+cw*.18,cy0+ch*.5),ch*.23,transform=fig.transFigure,facecolor='none',edgecolor='white' if cl!='MONITOREAR' else 'black',linewidth=1)); fig.text(cx0+cw*.62,cy0+ch*.65,'CONDICIÓN',fontsize=6.2,color=txt,ha='center',fontweight='bold'); fig.text(cx0+cw*.62,cy0+ch*.31,cl,fontsize=9,color=txt,ha='center',fontweight='bold')
    keys=list(data.keys())
    positions=[(.025,.515,.465,.285),(.510,.515,.465,.285),(.025,.215,.465,.285),(.510,.215,.465,.285)]
    for k,pos in zip(keys,positions): panel(k,data[k],*pos)
    # Footer simplified with no overlap
    fy=.065; fh=.125
    axsc=fig.add_axes([.025,fy,.29,fh]); axsc.axis('off'); axsc.add_patch(FancyBboxPatch((0,0),1,1,boxstyle='round,pad=0.006',facecolor='white',edgecolor='#d7e0ec')); axsc.text(.08,.87,'ESCALA DE ESPESOR',fontsize=8.5,color='#0b2545',fontweight='bold')
    cax=axsc.inset_axes([.07,.16,.055,.66]); grad=np.linspace(vmin,vmax,256).reshape(-1,1); cax.imshow(grad,cmap=cmap,norm=norm,origin='lower',aspect='auto'); cax.set_xticks([]); cax.set_yticks(np.linspace(0,255,5)); cax.set_yticklabels([f'{v:.2f}' for v in np.linspace(vmin,vmax,5)],fontsize=6.5); [sp.set_visible(False) for sp in cax.spines.values()]; axsc.text(.30,.70,'MAYOR ESPESOR',fontsize=7,color='#168323',fontweight='bold'); axsc.text(.30,.22,'MENOR ESPESOR',fontsize=7,color='red',fontweight='bold')
    axs=fig.add_axes([.325,fy,.185,fh]); axs.axis('off'); axs.add_patch(FancyBboxPatch((0,0),1,1,boxstyle='round,pad=0.006',facecolor='white',edgecolor='#d7e0ec')); axs.text(.12,.87,'SIMBOLOGÍA',fontsize=8.5,color='#0b2545',fontweight='bold')
    for i,(s,t) in enumerate([('○','Punto de medición'),('□','Menor espesor'),('◎','Centroide'),('—','Curva de nivel')]): axs.text(.12,.67-i*.17,s,fontsize=10,color='black'); axs.text(.26,.68-i*.17,t,fontsize=7,color='#0b2545')
    axr=fig.add_axes([.520,fy,.220,fh]); axr.axis('off'); axr.add_patch(FancyBboxPatch((0,0),1,1,boxstyle='round,pad=0.006',facecolor='white',edgecolor='#d7e0ec')); axr.text(.26,.87,'MAPA DE RIESGO (RESUMEN)',fontsize=8.5,color='#0b2545',fontweight='bold')
    for i,k in enumerate(keys): cl,co,_=cond(data[k]); xx=.13+i*.22; axr.text(xx,.58,k.replace(' ','\n'),fontsize=6.8,color='#0b2545',ha='center',fontweight='bold'); axr.scatter(xx,.25,s=300,color=co,edgecolor='black')
    axd=fig.add_axes([.750,fy,.225,fh]); axd.axis('off'); axd.add_patch(FancyBboxPatch((0,0),1,1,boxstyle='round,pad=0.006',facecolor='white',edgecolor='#d7e0ec')); axd.text(.20,.87,'DIAGNÓSTICO RESUMEN',fontsize=8.5,color='#0b2545',fontweight='bold')
    for i,k in enumerate(keys): cl,co,_=cond(data[k]); txt=f'{k}: '+('desgaste uniforme.' if cl=='NORMAL' else ('monitorear tendencia.' if cl=='MONITOREAR' else 'revisar condición.')); yy=.66-i*.18; axd.scatter(.08,yy,s=120,color=co,edgecolor='black'); axd.text(.15,yy,txt,fontsize=6.8,color='#0b2545',va='center')
    fig.text(.025,.025,'NOTA: Medidas en mm. Valores interpolados como apoyo visual; los puntos de medición representan los datos reales.',fontsize=7.3,color='#0b2545'); fig.text(.890,.025,'LUFUSSA   |   CONFIDENCIAL',fontsize=8,color='#65758b')
    return fig

st.markdown('<div class="box"><div class="title">OEM Insight Studio Web v1.3</div><div class="sub">Logo incluido, mejor espaciado y reporte sin textos montados.</div></div>', unsafe_allow_html=True)
st.info("Carga ambos Excel. Si las celdas son las mismas que el archivo original, usa 'Plantilla original'.")
c1,c2=st.columns(2)
with c1: file_a=st.file_uploader('Excel Banco A',type=['xlsx'],key='a')
with c2: file_b=st.file_uploader('Excel Banco B',type=['xlsx'],key='b')
st.subheader('Datos del reporte')
m1,m2,m3=st.columns(3)
with m1: unidad=st.text_input('Equipo / Unidad','GENSET 11'); fecha=st.text_input('Fecha',date.today().strftime('%d / %m / %Y'))
with m2: motor=st.text_input('Motor','WÄRTSILÄ 46'); reporte=st.text_input('Reporte No.','WSM-2026-0702')
with m3: horas=st.text_input('Horas de operación','12,845 h'); inspector=st.text_input('Inspector','')
if file_a and file_b:
    sa=sheets(file_a); sb=sheets(file_b); a,b=st.columns(2)
    with a: sheet_a=st.selectbox('Hoja Banco A',sa)
    with b: sheet_b=st.selectbox('Hoja Banco B',sb)
    mode=st.radio('Modo de lectura',['Plantilla original','Manual'],horizontal=True)
    if mode=='Manual':
        st.warning('Indica fila y columna de la esquina superior izquierda de cada matriz 5x3.')
        c1,c2,c3,c4=st.columns(4)
        with c1: au_r=st.number_input('A Upper fila',value=14,min_value=1); au_c=st.number_input('A Upper columna',value=7,min_value=1)
        with c2: al_r=st.number_input('A Lower fila',value=14,min_value=1); al_c=st.number_input('A Lower columna',value=18,min_value=1)
        with c3: bu_r=st.number_input('B Upper fila',value=14,min_value=1); bu_c=st.number_input('B Upper columna',value=7,min_value=1)
        with c4: bl_r=st.number_input('B Lower fila',value=14,min_value=1); bl_c=st.number_input('B Lower columna',value=18,min_value=1)
    cil_a=st.text_input('Nombre cilindro Banco A','A1'); cil_b=st.text_input('Nombre cilindro Banco B','B1')
    if st.button('Generar reporte',type='primary'):
        try:
            if mode=='Plantilla original':
                au=read_template(file_a,sheet_a,True); al=read_template(file_a,sheet_a,False); bu=read_template(file_b,sheet_b,True); bl=read_template(file_b,sheet_b,False)
            else:
                au=read_manual(file_a,sheet_a,int(au_r),int(au_c)); al=read_manual(file_a,sheet_a,int(al_r),int(al_c)); bu=read_manual(file_b,sheet_b,int(bu_r),int(bu_c)); bl=read_manual(file_b,sheet_b,int(bl_r),int(bl_c))
            data={f'{cil_a} UPPER':au,f'{cil_a} LOWER':al,f'{cil_b} UPPER':bu,f'{cil_b} LOWER':bl}; meta={'unidad':unidad,'motor':motor,'horas':horas,'fecha':fecha,'reporte':reporte,'inspector':inspector,'cil_a':cil_a,'cil_b':cil_b}
            fig=create_report(data,meta); png=BytesIO(); pdf=BytesIO(); fig.savefig(png,format='png',bbox_inches='tight',facecolor='white'); fig.savefig(pdf,format='pdf',bbox_inches='tight',facecolor='white'); png.seek(0); pdf.seek(0); plt.close(fig)
            st.success('Reporte generado correctamente.'); st.image(png,use_container_width=True); st.download_button('Descargar PNG',png.getvalue(),'Reporte_BEB_WSM.png','image/png'); st.download_button('Descargar PDF',pdf.getvalue(),'Reporte_BEB_WSM.pdf','application/pdf')
        except Exception as e: st.error(f'No se pudo generar el reporte: {e}')
else: st.warning('Carga ambos archivos Excel para continuar.')
